"""
PipeUpdUV - Salesforce Pipeline Data Integration Tool

This script integrates Salesforce pipeline export data (XLS/XLSX) into an existing Excel
tracking file (XLSM) for B2B sales opportunity management. The system focuses on
Opportunity (OpTY), Quotes, and Claims tracking while preserving manually entered data.

Version: 2.0000

Enhancement Summary:
- Added comprehensive logging system with rotating file handler
- Implemented custom exception classes for better error categorization
- Enhanced data validation and sanitization functions
- Added configuration validation on startup
- Optimized data processing for better performance
- Added type hints throughout the codebase for maintainability
- Improved error handling with try-catch blocks
- Enhanced documentation with detailed docstrings
Author: PipeUpdUV Team
Last Modified: 2025

Key Features:
- Enhanced error handling and logging
- Data validation and sanitization
- Performance optimizations
- Configuration validation
- Type hints for better code maintainability

Usage:
    python UpdatePipe.py [file_path|all]

    Arguments:
        file_path: Process specific pipe file
        all: Process all files in the pipe directory
        (no args): Process latest pipe file
"""

import math
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import numbers
from datetime import date,datetime
import pandas as pd
import openpyxl
import glob
import os
import warnings
import sys
import time
import re
import shutil
import logging
from typing import List, Optional, Tuple, Dict, Any
from pathlib import Path
from dotenv import load_dotenv
from colorama import Fore, Style, init

# Initialize colorama for Windows compatibility
init(autoreset=True)

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

################################################################
# Version Information
################################################################
VERSION_MAJOR = 2
VERSION_MINOR = "0000"
VERSION_STRING = f"{VERSION_MAJOR}.{VERSION_MINOR}"

# Excel format specifications
V1_COLUMN_COUNT = 21  # Original format (without Week columns)
V2_COLUMN_COUNT = 26  # Current format (with 5 Week columns)

# Custom formatter for colored DEBUG and ERROR messages
class ColoredFormatter(logging.Formatter):
    def format(self, record):
        # Get the standard formatted message
        formatted = super().format(record)

        # Apply colors to specific log levels after the timestamp
        if record.levelname == 'DEBUG':
            # Split at the first " - " after timestamp to preserve timestamp formatting
            parts = formatted.split(' - ', 2)
            if len(parts) >= 3:
                timestamp = parts[0]
                level = parts[1]
                message = parts[2]
                return f"{timestamp} - {Fore.YELLOW}{level} - {message}{Style.RESET_ALL}"
        elif record.levelname == 'WARNING':
            # Apply cyan color to WARNING messages (more visible on black backgrounds)
            parts = formatted.split(' - ', 2)
            if len(parts) >= 3:
                timestamp = parts[0]
                level = parts[1]
                message = parts[2]
                return f"{timestamp} - {Fore.CYAN}{level} - {message}{Style.RESET_ALL}"
        elif record.levelname == 'ERROR':
            # Apply red color to ERROR messages
            parts = formatted.split(' - ', 2)
            if len(parts) >= 3:
                timestamp = parts[0]
                level = parts[1]
                message = parts[2]
                return f"{timestamp} - {Fore.RED}{level} - {message}{Style.RESET_ALL}"

        return formatted

# Configure logging with proper encoding support and colored formatter
# Note: actual log level will be set after loading .env file
file_handler = logging.FileHandler('updatepipe.log', encoding='utf-8')
console_handler = logging.StreamHandler(sys.stdout)

# Apply custom formatter to console only (file should remain uncolored)
file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
console_handler.setFormatter(ColoredFormatter('%(asctime)s - %(levelname)s - %(message)s'))

logging.basicConfig(
    level=logging.INFO,  # This will be updated after loading .env
    handlers=[file_handler, console_handler]
)
logger = logging.getLogger(__name__)

load_dotenv()

# Configure logging level from environment
LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()
log_level_mapping = {
    'DEBUG': logging.DEBUG,
    'INFO': logging.INFO,
    'WARNING': logging.WARNING,
    'ERROR': logging.ERROR,
    'CRITICAL': logging.CRITICAL
}
actual_log_level = log_level_mapping.get(LOG_LEVEL, logging.INFO)

# Update logging level based on environment configuration
logger.setLevel(actual_log_level)
for handler in logging.getLogger().handlers:
    handler.setLevel(actual_log_level)

DIRECTORY_PIPE_RAW = os.getenv("DIRECTORY_PIPE_RAW")
INPUT_SUIVI_RAW = os.getenv("INPUT_SUIVI_RAW")
OUTPUT_SUIVI_RAW = os.getenv("OUTPUT_SUIVI_RAW")

SKIP_ROW = os.getenv("SKIP_ROW")

if (SKIP_ROW == None):
    SKIP_ROW=0
else:
    SKIP_ROW = int(SKIP_ROW)

GRANULARITE = os.getenv("GRANULARITE")
if (GRANULARITE == None): GRANULARITE='Date'

GRANULARITE_COL = os.getenv("GRANULARITE_COL")
if (GRANULARITE_COL == None):
    GRANULARITE_COL=0
else:
    GRANULARITE_COL = int(GRANULARITE_COL)

NORMAXDELTA = os.getenv("NORMAXDELTA")
if (NORMAXDELTA == None):
    NORMAXDELTA=10000000
else:
    NORMAXDELTA = int(NORMAXDELTA)

ROLLINGWINDOWS = os.getenv("ROLLINGWINDOWS")
if (ROLLINGWINDOWS == None):
    ROLLINGWINDOWS=31
else:
    ROLLINGWINDOWS = int(ROLLINGWINDOWS)

ROLLINGFIELD = os.getenv("ROLLINGFIELD")
if (ROLLINGFIELD == None): ROLLINGFIELD='Date'

BCKUP_PIPE_FILE = (str(os.getenv("BCKUP_PIPE_FILE")).lower() == 'true' )
if (BCKUP_PIPE_FILE == None): BCKUP_PIPE_FILE=False

if BCKUP_PIPE_FILE:
    BCKUP_DIRECTORY = os.getenv("BCKUP_DIRECTORY")
    if (BCKUP_DIRECTORY == None):
        BCKUP_PIPE_FILE=False
    else:
        BCKUP_PIPE_FILE = True

if BCKUP_PIPE_FILE:
    BCKUP_GRANULARITY = os.getenv("BCKUP_GRANULARITY")
    if (BCKUP_GRANULARITY == None): BCKUP_GRANULARITY="Days"

CURWEEK = os.getenv("CURWEEK")
if (CURWEEK != None):
    CURWEEK = int(CURWEEK)
else:
    CURWEEK = None

EXCLUDED_OPTY_OWNERS = os.getenv("EXCLUDED_OPTY_OWNERS", "")
# Parse comma-separated list and strip whitespace
if EXCLUDED_OPTY_OWNERS:
    EXCLUDED_OPTY_OWNERS = [owner.strip() for owner in EXCLUDED_OPTY_OWNERS.split(',') if owner.strip()]
else:
    EXCLUDED_OPTY_OWNERS = []

# Owner Opty Tracking: Starting line for Tab 2 (Last 5 Weeks Detail)
LINE_LAST_5W_OPTY = int(os.getenv("LINE_LAST_5W_OPTY", "28"))

# Hidden tabs: List of Excel sheet names to hide
HIDDEN_TABS = os.getenv("HIDDEN_TABS", "Owner Opty Tracking,Week History,Pipeline Close Lost")
# Parse comma-separated list and strip whitespace
if HIDDEN_TABS:
    HIDDEN_TABS = [tab.strip() for tab in HIDDEN_TABS.split(',') if tab.strip()]
else:
    HIDDEN_TABS = []

# To avoid localisation colision
# Define col index for labels in Pipe file
# Only done for col name with problem
COL_OPTYOWNER=0
COL_CREATED=1
COL_CLOSED=2
COL_STAGE=3
COL_CUSTOMER=6
COL_TOTPRICE=9
COL_SALESMODELNAME=10

# Global variable to store actual column names (language-specific)
# This is set during dataframe processing after column reorganization
# Allows language-independent access via cols[COL_*] pattern
cols = None

################################################################
# Exception Classes
################################################################
class PipeProcessingError(Exception):
    """Custom exception for pipe processing errors"""
    pass

class ConfigurationError(Exception):
    """Custom exception for configuration errors"""
    pass

class DataValidationError(Exception):
    """Custom exception for data validation errors"""
    pass

################################################################
# Configuration Validation
################################################################
def display_environment_config() -> None:
    """Display all environment variables and their values in DEBUG mode"""
    if logger.getEffectiveLevel() <= logging.DEBUG:
        logger.debug("="*60)
        logger.debug("PIPEUPDUV CONFIGURATION")
        logger.debug("="*60)

        # Version information
        logger.debug(f"VERSION = {VERSION_STRING} (Major: {VERSION_MAJOR}, Minor: {VERSION_MINOR})")
        logger.debug(f"Excel V1 format = {V1_COLUMN_COUNT} columns")
        logger.debug(f"Excel V2 format = {V2_COLUMN_COUNT} columns")
        logger.debug("-" * 60)

        # Core configuration
        logger.debug(f"DIRECTORY_PIPE_RAW = {repr(DIRECTORY_PIPE_RAW)}")
        logger.debug(f"INPUT_SUIVI_RAW = {repr(INPUT_SUIVI_RAW)}")
        logger.debug(f"OUTPUT_SUIVI_RAW = {repr(OUTPUT_SUIVI_RAW)}")

        # Data processing configuration
        logger.debug(f"SKIP_ROW = {SKIP_ROW} (default: 12)")
        logger.debug(f"GRANULARITE = {repr(GRANULARITE)} (default: 'Date')")
        logger.debug(f"GRANULARITE_COL = {GRANULARITE_COL} (default: 0)")

        # Analysis configuration
        logger.debug(f"NORMAXDELTA = {NORMAXDELTA} (default: 10000000)")
        logger.debug(f"ROLLINGWINDOWS = {ROLLINGWINDOWS} (default: 31)")
        logger.debug(f"ROLLINGFIELD = {repr(ROLLINGFIELD)} (default: 'Date')")

        # Testing configuration
        logger.debug(f"CURWEEK = {CURWEEK} (default: None - use current week)")

        # Owner Opportunity Tracking configuration
        logger.debug(f"EXCLUDED_OPTY_OWNERS = {EXCLUDED_OPTY_OWNERS} (default: [])")
        logger.debug(f"LINE_LAST_5W_OPTY = {LINE_LAST_5W_OPTY} (default: 28)")

        # Excel tab configuration
        logger.debug(f"HIDDEN_TABS = {HIDDEN_TABS} (default: ['Owner Opty Tracking', 'Week History', 'Pipeline Close Lost'])")

        # Backup configuration
        logger.debug(f"BCKUP_PIPE_FILE = {BCKUP_PIPE_FILE} (default: False)")
        if BCKUP_PIPE_FILE:
            logger.debug(f"BCKUP_DIRECTORY = {repr(globals().get('BCKUP_DIRECTORY', 'Not set'))}")
            logger.debug(f"BCKUP_GRANULARITY = {repr(globals().get('BCKUP_GRANULARITY', 'days'))}")

        # Logging configuration
        logger.debug(f"LOG_LEVEL = {repr(LOG_LEVEL)} (default: 'INFO')")
        logger.debug(f"Effective log level = {logging.getLevelName(logger.getEffectiveLevel())}")

        logger.debug("="*60)

def validate_configuration() -> None:
    """Validate all required configuration parameters"""
    required_configs = {
        'DIRECTORY_PIPE_RAW': DIRECTORY_PIPE_RAW,
        'INPUT_SUIVI_RAW': INPUT_SUIVI_RAW,
        'OUTPUT_SUIVI_RAW': OUTPUT_SUIVI_RAW
    }

    missing_configs = [key for key, value in required_configs.items() if not value]
    if missing_configs:
        raise ConfigurationError(f"Missing required configuration: {', '.join(missing_configs)}")

    # Validate paths exist
    if not os.path.exists(DIRECTORY_PIPE_RAW):
        raise ConfigurationError(f"Pipe directory does not exist: {DIRECTORY_PIPE_RAW}")

    if not os.path.exists(INPUT_SUIVI_RAW):
        raise ConfigurationError(f"Input tracking file does not exist: {INPUT_SUIVI_RAW}")

    # Validate numeric configurations
    if ROLLINGWINDOWS > 31:
        logger.warning(f"ROLLINGWINDOWS value {ROLLINGWINDOWS} exceeds recommended maximum of 31")

    logger.info("Configuration validation completed successfully")

################################################################
# Functions Helper
################################################################
def BackupPipeBefore(pipefullpath: str) -> bool:

    now = datetime.now()

    # dd/mm/YY
    dtstr = now.strftime("%Y%m%d-%H")

    filename = os.path.basename(pipefullpath).split('.')
    if len(filename) == 2:
        targetbckfn = f'{BCKUP_DIRECTORY}\\{filename[0]}-{dtstr}-bck.{filename[1]}'
    else:
        logger.error(f'Error building backup filename from {pipefullpath}')
        return False

    # Search on the exact name or on a part including only the day
    if BCKUP_GRANULARITY.lower() == "days":
        Targetsrch = f'{BCKUP_DIRECTORY}\\{filename[0]}-{now.strftime("%Y%m%d")}*-bck.{filename[1]}'
    else:
        Targetsrch = targetbckfn

    files = glob.glob(Targetsrch)

    if len(files) == 0:
        # Backup Pipe File
        try:
            shutil.copy(pipefullpath, targetbckfn)
            logger.info(f'Backup file created: {targetbckfn}')
        except Exception as e:
            logger.error(f'Failed to create backup: {str(e)}')
            return False

    return True

def GetLatestPipe(idir: str) -> str:
    """Get the latest pipe file from directory"""
    try:
        files = glob.glob(f'{idir}/*.xls*')
        if not files:
            raise PipeProcessingError(f"No Excel files found in directory: {idir}")

        latest_file = max(files, key=os.path.getctime)
        # Create colored debug message for latest pipe file
        filename = os.path.basename(latest_file)
        colored_latest_message = f"Latest pipe file found: {Fore.GREEN}{filename}{Style.RESET_ALL}"
        logger.debug(colored_latest_message)
        return latest_file
    except Exception as e:
        logger.error(f"Error finding latest pipe file: {str(e)}")
        raise PipeProcessingError(f"Failed to find latest pipe file: {str(e)}")

def GetAllPipe(idir: str) -> List[str]:
    """Get all pipe files from directory, sorted by creation time"""
    try:
        files = glob.glob(f'{idir}/*.xls*')
        if not files:
            raise PipeProcessingError(f"No Excel files found in directory: {idir}")

        files.sort(key=os.path.getctime)
        logger.info(f"Found {len(files)} pipe files")
        return files
    except Exception as e:
        logger.error(f"Error getting all pipe files: {str(e)}")
        raise PipeProcessingError(f"Failed to get pipe files: {str(e)}")
def DetectHeaderRow(pfile: str, max_rows: int = 30) -> int:
    """Automatically detect the header row in Salesforce extract files

    Salesforce extracts may have varying numbers of header/warning lines.
    This function scans the first rows to find the actual column headers.
    Supports multiple languages (English, French, etc.)

    Args:
        pfile: Path to the Excel file
        max_rows: Maximum number of rows to scan (default 30)

    Returns:
        Row number where headers are found (0-based index for skiprows parameter)
        Returns 0 if headers are found immediately or if detection fails

    Raises:
        PipeProcessingError: If header row cannot be found
    """
    try:
        # Read first N rows without headers
        df = pd.read_excel(pfile, nrows=max_rows, header=None)

        # Known header columns that should always be present in Salesforce extracts
        # Multiple language variants: each inner list is a language set
        # We need at least 2 matches from ANY language set
        expected_headers_multilang = [
            # English
            ['Opportunity Owner', 'Created Date', 'Close Date', 'Stage'],
            # French
            ['Propriétaire de l\'opportunité', 'Date de création', 'Date de clôture', 'Étape'],
        ]

        # Scan each row looking for these headers
        for row_idx in range(len(df)):
            row = df.iloc[row_idx]
            # Convert row values to strings and check for expected headers
            row_str = row.astype(str).str.strip()

            # Check each language set
            for lang_idx, expected_headers in enumerate(expected_headers_multilang):
                # Count how many expected headers are found in this row
                matches = sum(1 for header in expected_headers if header in row_str.values)

                # If we find at least 2 of the expected headers, this is our header row
                if matches >= 2:
                    lang_name = ['English', 'French'][lang_idx] if lang_idx < 2 else f'Language {lang_idx}'
                    logger.info(f"Auto-detected header row at line {row_idx + 1} (will skip {row_idx} rows) - {lang_name} format")
                    logger.debug(f"Found {matches} matching headers: {[h for h in expected_headers if h in row_str.values]}")
                    return row_idx

        # If no header found in first max_rows, log warning and use SKIP_ROW or default
        if SKIP_ROW > 0:
            logger.warning(f"Could not auto-detect header row in first {max_rows} rows, using SKIP_ROW={SKIP_ROW}")
            return SKIP_ROW
        else:
            logger.warning(f"Could not auto-detect header row in first {max_rows} rows, defaulting to row 12")
            return 12

    except Exception as e:
        error_msg = f"Error during header detection: {str(e)}"
        logger.error(error_msg)
        if SKIP_ROW > 0:
            logger.warning(f"Falling back to SKIP_ROW={SKIP_ROW}")
            return SKIP_ROW
        raise PipeProcessingError(error_msg)

def CheckPipeFile(pfile: str) -> bool:
    """Check if pipe file is valid Excel file"""
    try:
        if not os.path.isfile(pfile):
            logger.error(f"File does not exist: {pfile}")
            return False

        ext = os.path.splitext(pfile)[-1].lower()
        if ext not in ['.xls', '.xlsx']:
            logger.error(f"Invalid file extension: {ext}. Expected .xls or .xlsx")
            return False

        # Try to read the file to ensure it's not corrupted
        try:
            pd.read_excel(pfile, nrows=1)
        except Exception as e:
            logger.error(f"File appears to be corrupted: {str(e)}")
            return False

        logger.debug(f"Pipe file validation successful: {pfile}")
        return True
    except Exception as e:
        logger.error(f"Error validating pipe file {pfile}: {str(e)}")
        return False

################################################################
# Data Validation Functions
################################################################
def validate_dataframe_structure(df: pd.DataFrame, expected_cols: List[str], df_name: str) -> None:
    """Validate that dataframe has expected structure"""
    if df.empty:
        raise DataValidationError(f"{df_name} is empty")

    missing_cols = set(expected_cols) - set(df.columns)
    if missing_cols:
        logger.warning(f"{df_name} missing expected columns: {missing_cols}")

def sanitize_numeric_value(value: Any, default: float = 0.0) -> float:
    """Safely convert value to numeric, with fallback"""
    try:
        if pd.isna(value) or value == '':
            return default
        # Remove currency symbols and formatting
        if isinstance(value, str):
            cleaned = re.sub(r'[^\d.-]', '', str(value))
            return float(cleaned) if cleaned else default
        return float(value)
    except (ValueError, TypeError):
        logger.warning(f"Could not convert '{value}' to numeric, using default {default}")
        return default

def sanitize_date_value(value: Any) -> Optional[datetime]:
    """Safely convert value to datetime"""
    try:
        if pd.isna(value) or value == '':
            return None
        return pd.to_datetime(value, format='mixed')
    except Exception as e:
        logger.warning(f"Could not convert '{value}' to datetime: {str(e)}")
        return None

def sanitize_string_value(value: Any, default: str = '') -> str:
    """Safely convert value to string"""
    try:
        if pd.isna(value):
            return default
        return str(value).strip()
    except Exception:
        return default

#Mapping Date to Quarter FYear
def GetQFFromDate(cdate: datetime) -> Tuple[int, str]:

    cm = cdate.month
    if cm < 4:
        Quarter = 1
    else:
        if cm < 7:
            Quarter = 2
        else:
            if cm < 10:
                Quarter = 3
            else:
                Quarter = 4

    Year = str(cdate.year)[-2:]

    return Quarter, Year


#Generic Mapping Functions
def Mapping_Generic(Key: str, Col: str) -> str:
    """Generic mapping function to get value from master dataframe"""
    try:
        if df_master is None or df_master.empty:
            return ''

        rowval = df_master.loc[df_master['Key'] == Key]
        if len(rowval) == 0:
            return ''

        rtv = rowval.at[rowval.index[-1], Col]
        return sanitize_string_value(rtv)
    except Exception as e:
        logger.debug(f"Error in Mapping_Generic for Key {Key}, Col {Col}: {str(e)}")
        return ''

#Mapping Functions for
# 'Estimated\nQuantity', 'Revenu From\nEstinated Qty', 'Quarter Invoice\nFacturation', 'Forecast projet\nMenu déroulant', 'Next Step & Support demandé / Commentaire'

def Mapping_Qty(Key: str) -> Any:
    """Map quantity values with formula handling"""
    try:
        eq = Mapping_Generic(Key, 'Estimated\nQuantity')

        if str(eq).startswith('=') or str(eq) == '':
            rowval = df_master.loc[df_master['Key'] == Key]
            if len(rowval) > 0 and 'Quantité' in rowval.columns:
                eq = rowval['Quantité'].values[0]

        return sanitize_numeric_value(eq) if eq else ''
    except Exception as e:
        logger.debug(f"Error in Mapping_Qty for Key {Key}: {str(e)}")
        return ''

def Mapping_RevEur(Key: str) -> Any:
    """Map revenue values with formula and calculation handling"""
    try:
        rev = Mapping_Generic(Key, 'Revenu From\nEstinated Qty')

        if rev and rev != '':
            rowval = df_master.loc[df_master['Key'] == Key]
            if len(rowval) == 0:
                return ''

            if str(rev).startswith('='):
                if 'Prix total' in rowval.columns:
                    rev = sanitize_numeric_value(rowval['Prix total'].values[0])
            else:
                # Calculate from quantity and price
                if 'Estimated\nQuantity' in rowval.columns and 'Prix de vente' in rowval.columns:
                    qty = sanitize_numeric_value(rowval['Estimated\nQuantity'].values[0])
                    price = sanitize_numeric_value(rowval['Prix de vente'].values[0])
                    rev = qty * price

        return rev if rev else ''
    except Exception as e:
        logger.debug(f"Error in Mapping_RevEur for Key {Key}: {str(e)}")
        return ''

def Mapping_QtrInvoice(Key: str) -> str:
    """Map quarter invoice values with automatic calculation from close date

    Args:
        Key: Unique key for the opportunity

    Returns:
        Quarter invoice string in format QnFYyy or preserved value
    """

     # Rules :
     # If nothing, leave nothing
     # if lenght of value is not 2,4 or 6 put nothing
     # If first letter of value is Q, get the close date and calculate the QnFYyy

    eq = Mapping_Generic(Key,'Quarter Invoice\nFacturation')

    seq = str(eq)

    # Update : We translate the Close Date in QnFy even if the field is blank - Then it can eventually be changed. As long as it is in the  right format this wont be changed here.
    try:

        CloseDate = Mapping_Generic(Key,'Date de clôture')
        if str(CloseDate) != '':
            Quarter,Year = GetQFFromDate(CloseDate)
            seq = f'Q{Quarter}FY{Year}'

    except:
        pass


    return seq

def Mapping_FrCast(row: pd.Series) -> str:
    """Map forecast values based on Win Rate and Stage

    Args:
        row: Pandas Series containing opportunity data

    Returns:
        Forecast category string
    """
    global cols

    Key = row['Key']

    eq = Mapping_Generic(Key,'Forecast projet\nMenu déroulant')
    seq = str(eq)

    AS = ["LOST = Perdu", "UNCOMMITED = Pas certain", "UNCOMMITED UPSIDE = Certain à 50% du WIN","COMMIT AT RISK = Certain à 75% du WIN","COMMIT = Certain à 100% du WIN","WIN = Gagné"]

# Update : Automatic fill of the column value base on Win Rate column ... If not empty
    fcast = seq

    # Use language-independent column access via cols[COL_STAGE]
    # This works with both English "Stage" and French "Étape" columns
    Stat = row[cols[COL_STAGE]] if cols is not None else row.get('Stage', '')

    if Stat.lower() == 'closed won':
        fcast = "WIN = Gagné"
    else:
        if seq not in AS:
            try:
                # rowval = df_master.loc[df_master['Key'] == Key]
                # WR = rowval['Win Rate'].values[0].replace('%', '')
                WR = row['Win Rate'].replace('%', '')
                # Check if the value is NaN or space (after converting to string and stripping whitespace)
                if pd.isna(WR) or str(WR).strip() == '':
                    return ''
                
                # Convert the value to a float to handle numeric comparison
                WR = float(WR)
                
                # Number of ranges is the same as the length of the AS array
                num_ranges = len(AS)
                range_size = 100 / num_ranges
                
                # Find the index in the AS array that corresponds to the value
                index = min(int((WR - 1) / range_size), num_ranges - 1)
                fcast = AS[index]

            except:
                pass

    return fcast

def Mapping_NxtStp(Key: str) -> str:
    """Map next step comments from existing data

    Args:
        Key: Unique key for the opportunity

    Returns:
        Next step comment string
    """

    return Mapping_Generic(Key,'Next Step & Support demandé / Commentaire')

def DetectWeekShift(df_master: pd.DataFrame) -> Tuple[int, List[str]]:
    """Detect if there will be a week shift and return shift amount and existing week columns

    Args:
        df_master: Master DataFrame containing existing week columns

    Returns:
        Tuple of (shift_amount, existing_week_columns)
    """
    try:
        current_week = datetime.now().isocalendar()[1] if CURWEEK is None else CURWEEK

        # Find existing week columns in order (should be columns V-Z, positions 22-26)
        existing_week_columns = []
        for col in df_master.columns:
            if col and str(col).startswith('Week '):
                existing_week_columns.append(str(col))

        if not existing_week_columns:
            logger.info("No existing week columns found, no shift needed")
            return 0, []

        # Find the center column (should be column X, the 3rd of 5 columns)
        if len(existing_week_columns) >= 3:
            center_column = existing_week_columns[2]  # Column X (0-indexed: V=0, W=1, X=2)

            # Extract week number from center column (e.g., "Week 39" -> 39)
            if center_column.startswith('Week '):
                try:
                    center_week = int(center_column.replace('Week ', ''))
                    shift_amount = current_week - center_week

                    # Handle year rollover
                    if shift_amount > 26:  # More than half year forward
                        shift_amount -= 52  # Assume we went to previous year
                    elif shift_amount < -26:  # More than half year backward
                        shift_amount += 52  # Assume we went to next year

                    logger.info(f"Detected week shift: current week {current_week}, center was {center_week}, shift amount: {shift_amount}")
                    return shift_amount, existing_week_columns
                except ValueError:
                    logger.warning(f"Could not parse week number from {center_column}")

        logger.info("Could not determine week shift, assuming no shift")
        return 0, existing_week_columns

    except Exception as e:
        logger.error(f"Error detecting week shift: {str(e)}")
        return 0, []

def ApplyWeekShiftFromHistory(df_master: pd.DataFrame, df_whisto: pd.DataFrame, new_week_columns: List[str]) -> pd.DataFrame:
    """Apply week shift using data from Week History DataFrame

    Args:
        df_master: Master DataFrame to update
        df_whisto: Week History DataFrame containing historical data
        new_week_columns: List of new week column names (e.g., ['Week 39', 'Week 40', ...])

    Returns:
        Updated DataFrame with week data from history
    """
    try:
        if df_whisto.empty or not new_week_columns:
            logger.info("No week shift to apply - empty history or no columns")
            return df_master

        logger.info(f"Applying week shift using history data for columns: {new_week_columns}")

        # Find existing week columns in df_master
        existing_week_columns = [col for col in df_master.columns if col and str(col).startswith('Week ')]

        if len(existing_week_columns) < 5:
            logger.warning(f"Expected 5 week columns, found {len(existing_week_columns)}")
            return df_master

        # For each row in df_master that has a Key
        for idx, row in df_master.iterrows():
            if 'Key' not in row or pd.isna(row['Key']) or str(row['Key']).strip() == '':
                continue

            key = str(row['Key'])

            # Find this key in the history DataFrame
            history_rows = df_whisto[df_whisto['key'] == key]
            if len(history_rows) == 0:
                # Clear week columns for this key since no history exists
                for col in existing_week_columns:
                    if col in df_master.columns:
                        df_master.at[idx, col] = ''
                continue

            history_row = history_rows.iloc[0]

            # For each of the 5 week columns in the new structure
            for col_idx, new_week_col in enumerate(new_week_columns):
                if col_idx >= len(existing_week_columns):
                    break  # Safety check

                target_col = existing_week_columns[col_idx]  # V, W, X, Y, Z positions

                # Extract week number from new_week_col (e.g., "Week 39" -> 39)
                try:
                    week_num = int(new_week_col.replace('Week ', ''))
                    history_col = f'W{week_num:02d}'

                    # Get value from history DataFrame
                    if history_col in history_row and pd.notna(history_row[history_col]):
                        value = str(history_row[history_col]).strip()
                        df_master.at[idx, target_col] = value if value != '' else ''
                    else:
                        df_master.at[idx, target_col] = ''

                    logger.debug(f"Key {key}: Position {col_idx} ({target_col}) gets {new_week_col} data from {history_col} = '{value if 'value' in locals() else ''}'")

                except ValueError:
                    logger.warning(f"Could not parse week number from {new_week_col}")
                    df_master.at[idx, target_col] = ''

        logger.info(f"Week shift from history completed")
        return df_master

    except Exception as e:
        logger.error(f"Error applying week shift from history: {str(e)}")
        return df_master

def GetDynamicWeekColumns() -> List[str]:
    """Generate the 5 dynamic week column names based on current week

    Returns:
        List of 5 week column names: [Week-2, Week-1, Week, Week+1, Week+2]
    """
    if CURWEEK is not None:
        current_week = CURWEEK
        logger.debug(f'Using test week number: {current_week}')  # Only log at debug level to reduce noise
    else:
        current_week = datetime.now().isocalendar()[1]
        # Don't log here, let the calling code log the week columns being used

    week_columns = []

    for offset in range(-2, 3):  # -2, -1, 0, +1, +2
        week_num = current_week + offset
        # Handle year boundaries - most years have 52 weeks, some have 53
        if week_num < 1:
            # Get the actual number of weeks in the previous year
            prev_year = datetime.now().year - 1
            last_week_prev_year = datetime(prev_year, 12, 28).isocalendar()[1]  # Week containing Dec 28 is always the last week
            week_num = last_week_prev_year + week_num
        elif week_num > 52:
            # Check if current year actually has 53 weeks
            current_year = datetime.now().year
            last_week_current_year = datetime(current_year, 12, 28).isocalendar()[1]
            if week_num > last_week_current_year:
                week_num = week_num - last_week_current_year

        week_columns.append(f"Week {week_num}")

    return week_columns

def CreateWeekHistoryDataFrame() -> pd.DataFrame:
    """Create a new Week History DataFrame with proper column structure

    Returns:
        DataFrame with 'key' (for internal processing), 'Opportunity Number',
        'Model Name', and W01-W53 columns
    """
    columns = ['key', 'Opportunity Number', 'Model Name'] + [f'W{i:02d}' for i in range(1, 54)]  # W01 to W53
    return pd.DataFrame(columns=columns)

def LoadWeekHistoryFromExcel(workbook: openpyxl.Workbook) -> pd.DataFrame:
    """Load Week History data from Excel tab if it exists

    Supports both old format (with 'key' column) and new format
    (with 'Opportunity Number' and 'Model Name' columns).

    Args:
        workbook: Excel workbook to read from

    Returns:
        DataFrame containing Week History data or empty DataFrame with proper structure
    """
    try:
        if "Week History" in workbook.sheetnames:
            ws_whisto = workbook['Week History']
            # Convert to DataFrame
            df_whisto = pd.DataFrame(ws_whisto.values)
            if not df_whisto.empty:
                # Set column names from first row
                df_whisto.columns = df_whisto.iloc[0]
                df_whisto = df_whisto.drop(df_whisto.index[0]).reset_index(drop=True)

                # Check if this is old format (key only) or new format (Opportunity Number + Model Name)
                has_opty_model = 'Opportunity Number' in df_whisto.columns and 'Model Name' in df_whisto.columns
                has_key = 'key' in df_whisto.columns

                if has_opty_model:
                    # New format - ensure all columns exist
                    expected_columns = ['Opportunity Number', 'Model Name'] + [f'W{i:02d}' for i in range(1, 54)]
                    # Create 'key' column for internal processing (Opty Number + Model Name)
                    df_whisto['key'] = df_whisto['Opportunity Number'].astype(str) + df_whisto['Model Name'].astype(str)
                    # Add 'key' at the beginning for internal processing
                    expected_columns = ['key'] + expected_columns
                elif has_key:
                    # Old format - migrate to new format
                    logger.info("Migrating old Week History format (key) to new format (Opportunity Number + Model Name)")
                    # For old format, we can't split the key reliably, so create empty columns
                    # They will be populated on next update
                    df_whisto['Opportunity Number'] = ''
                    df_whisto['Model Name'] = ''
                    expected_columns = ['key', 'Opportunity Number', 'Model Name'] + [f'W{i:02d}' for i in range(1, 54)]
                else:
                    # Neither format - create new structure
                    logger.warning("Week History has unexpected format, creating new DataFrame")
                    return CreateWeekHistoryDataFrame()

                # Ensure we have all required columns
                for col in expected_columns:
                    if col not in df_whisto.columns:
                        df_whisto[col] = ''

                # Reorder columns to match expected structure
                df_whisto = df_whisto.reindex(columns=expected_columns, fill_value='')
                logger.info(f"Loaded Week History with {len(df_whisto)} rows")
                return df_whisto

        # If tab doesn't exist or is empty, create new DataFrame
        logger.info("Week History tab not found or empty, creating new DataFrame")
        return CreateWeekHistoryDataFrame()

    except Exception as e:
        logger.warning(f"Error loading Week History: {str(e)}, creating new DataFrame")
        return CreateWeekHistoryDataFrame()

def UpdateWeekHistoryRow(df_whisto: pd.DataFrame, key: str, week_data: Dict[str, str],
                         opty_number: str = '', model_name: str = '') -> pd.DataFrame:
    """Update or create a row in the Week History DataFrame

    Args:
        df_whisto: Week History DataFrame
        key: Unique key for the opportunity (Opty Number + Model Name)
        week_data: Dictionary mapping week column names to values
        opty_number: Opportunity Number (optional, for new rows)
        model_name: Model Name (optional, for new rows)

    Returns:
        Updated Week History DataFrame
    """
    try:
        # Find existing row with this key
        mask = df_whisto['key'] == key
        existing_rows = df_whisto[mask]

        if len(existing_rows) > 0:
            # Update existing row
            row_idx = existing_rows.index[0]
            # Update Opportunity Number and Model Name if provided and currently empty
            if opty_number and (pd.isna(df_whisto.at[row_idx, 'Opportunity Number']) or
                               str(df_whisto.at[row_idx, 'Opportunity Number']).strip() == ''):
                df_whisto.at[row_idx, 'Opportunity Number'] = opty_number
            if model_name and (pd.isna(df_whisto.at[row_idx, 'Model Name']) or
                              str(df_whisto.at[row_idx, 'Model Name']).strip() == ''):
                df_whisto.at[row_idx, 'Model Name'] = model_name
            # Update week data
            for week_col, value in week_data.items():
                # Convert week column name (e.g., "Week 25") to Week History format (e.g., "W25")
                if week_col.startswith('Week '):
                    week_num = week_col.replace('Week ', '')
                    whisto_col = f'W{int(week_num):02d}'
                    if whisto_col in df_whisto.columns:
                        df_whisto.at[row_idx, whisto_col] = value
        else:
            # Create new row
            new_row = {
                'key': key,
                'Opportunity Number': opty_number,
                'Model Name': model_name
            }
            # Initialize all week columns to empty
            for i in range(1, 54):
                new_row[f'W{i:02d}'] = ''

            # Fill in the provided week data
            for week_col, value in week_data.items():
                if week_col.startswith('Week '):
                    week_num = week_col.replace('Week ', '')
                    whisto_col = f'W{int(week_num):02d}'
                    if whisto_col in new_row:
                        new_row[whisto_col] = value

            # Add new row to DataFrame
            df_whisto = pd.concat([df_whisto, pd.DataFrame([new_row])], ignore_index=True)

        return df_whisto

    except Exception as e:
        logger.error(f"Error updating Week History row for key {key}: {str(e)}")
        return df_whisto

def CleanWeekHistory(df_whisto: pd.DataFrame, df_pipe: pd.DataFrame) -> pd.DataFrame:
    """Remove rows from Week History that no longer have corresponding keys in df_pipe

    Args:
        df_whisto: Week History DataFrame
        df_pipe: Current pipeline DataFrame with Key column

    Returns:
        Cleaned Week History DataFrame
    """
    try:
        if df_whisto.empty or df_pipe.empty:
            return df_whisto

        # Get list of current keys from df_pipe
        current_keys = set(df_pipe['Key'].tolist())

        # Filter df_whisto to keep only rows with keys that still exist in df_pipe
        initial_count = len(df_whisto)
        df_whisto_cleaned = df_whisto[df_whisto['key'].isin(current_keys)].copy()
        cleaned_count = len(df_whisto_cleaned)

        removed_count = initial_count - cleaned_count
        if removed_count > 0:
            logger.info(f"Cleaned Week History: removed {removed_count} orphaned rows, keeping {cleaned_count} rows")

        return df_whisto_cleaned

    except Exception as e:
        logger.error(f"Error cleaning Week History: {str(e)}")
        return df_whisto

def UpgradeFormatV1toV2(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Upgrade Excel format from V1 (21 columns) to V2 (26 columns) by adding Week columns

    Args:
        worksheet: Excel worksheet to upgrade
    """
    try:
        logger.info("Upgrading Excel format from V1 (21 columns) to V2 (26 columns)")

        # V2 adds 5 Week columns at the end (columns V, W, X, Y, Z)
        # Generate current week column names
        dynamic_week_columns = GetDynamicWeekColumns()

        # Find the last column with data
        max_col = worksheet.max_column

        # Add headers for the 5 new Week columns (row 2 is the header row)
        for i, week_col_name in enumerate(dynamic_week_columns):
            col_idx = max_col + 1 + i  # Add after existing columns
            worksheet.cell(row=2, column=col_idx).value = week_col_name  # Header row 2

        # Initialize empty values for all data rows for the new columns
        for row_num in range(3, worksheet.max_row + 1):  # Start from row 3 (data rows)
            for i in range(5):  # 5 new columns
                col_idx = max_col + 1 + i
                worksheet.cell(row=row_num, column=col_idx).value = ""

        logger.info(f"Successfully upgraded Excel format: added {len(dynamic_week_columns)} Week columns")

    except Exception as e:
        logger.error(f"Error upgrading Excel format V1 to V2: {str(e)}")
        raise PipeProcessingError(f"Failed to upgrade Excel format: {str(e)}")

def WriteWeekHistoryToExcel(workbook: openpyxl.Workbook, df_whisto: pd.DataFrame) -> None:
    """Write Week History DataFrame to Excel, replacing existing tab

    Writes Opportunity Number, Model Name, and week columns (W01-W53).
    The internal 'key' column is excluded from Excel output.

    Args:
        workbook: Excel workbook to write to
        df_whisto: Week History DataFrame to write
    """
    try:
        # Remove existing Week History tab if it exists
        if "Week History" in workbook.sheetnames:
            del workbook["Week History"]
            logger.debug("Removed existing Week History tab")

        # Create new Week History tab
        ws_whisto = workbook.create_sheet("Week History")

        # Create a copy of the DataFrame excluding the 'key' column for Excel output
        # Keep only: Opportunity Number, Model Name, and W01-W53 columns
        output_columns = ['Opportunity Number', 'Model Name'] + [f'W{i:02d}' for i in range(1, 54)]
        df_output = df_whisto[output_columns].copy()

        # Write data to the sheet
        for r in dataframe_to_rows(df_output, index=False, header=True):
            ws_whisto.append(r)

        logger.info(f"Written Week History with {len(df_output)} rows to Excel")

    except Exception as e:
        logger.error(f"Error writing Week History to Excel: {str(e)}")

################################################################
# Owner Opportunity Tracking Functions
################################################################

def CreateOwnerOpptyTrackingDataFrame() -> pd.DataFrame:
    """Create a new Owner Opportunity Tracking DataFrame with proper column structure

    Returns:
        DataFrame with 'owner' column and W01-W53 columns
    """
    columns = ['owner'] + [f'W{i:02d}' for i in range(1, 54)]  # W01 to W53
    return pd.DataFrame(columns=columns)

def LoadOwnerOpptyTrackingFromExcel(workbook: openpyxl.Workbook) -> pd.DataFrame:
    """Load Owner Opportunity Tracking data from Excel tab if it exists

    IMPORTANT: This function only reads Table 1 (summary counts), not Table 2 (details).
    The two tables are separated by empty rows in the Excel sheet.

    Args:
        workbook: Excel workbook to read from

    Returns:
        DataFrame containing Owner Opportunity Tracking data or empty DataFrame with proper structure
    """
    try:
        if "Owner Opty Tracking" in workbook.sheetnames:
            ws_otrack = workbook['Owner Opty Tracking']

            # Read only until we hit an empty row (separator between Table 1 and Table 2)
            # First row should be headers
            data_rows = []
            for row_idx, row in enumerate(ws_otrack.iter_rows(values_only=True)):
                # Skip if first cell (owner column) is None or empty
                # This stops at the separator rows between Table 1 and Table 2
                if row_idx == 0:
                    # First row is header
                    header_row = row
                    continue
                if row[0] is None or str(row[0]).strip() == '':
                    # Empty row found - this is the separator, stop reading Table 1
                    logger.debug(f"Stopped reading Owner Opty Tracking at row {row_idx + 1} (empty separator)")
                    break
                data_rows.append(row)

            if not data_rows:
                logger.info("Owner Opty Tracking tab exists but Table 1 is empty, creating new DataFrame")
                return CreateOwnerOpptyTrackingDataFrame()

            # Create DataFrame from only Table 1 data
            df_otrack = pd.DataFrame(data_rows, columns=header_row)

            # Ensure we have all required columns
            expected_columns = ['owner'] + [f'W{i:02d}' for i in range(1, 54)]
            for col in expected_columns:
                if col not in df_otrack.columns:
                    df_otrack[col] = 0

            # Reorder columns to match expected structure
            df_otrack = df_otrack.reindex(columns=expected_columns, fill_value=0)

            # Convert week columns to numeric
            for col in [f'W{i:02d}' for i in range(1, 54)]:
                df_otrack[col] = pd.to_numeric(df_otrack[col], errors='coerce').fillna(0).astype(int)

            logger.info(f"Loaded Owner Opty Tracking Table 1 with {len(df_otrack)} rows")
            return df_otrack

        # If tab doesn't exist or is empty, create new DataFrame
        logger.info("Owner Opty Tracking tab not found or empty, creating new DataFrame")
        return CreateOwnerOpptyTrackingDataFrame()

    except Exception as e:
        logger.warning(f"Error loading Owner Opty Tracking: {str(e)}, creating new DataFrame")
        return CreateOwnerOpptyTrackingDataFrame()

def ExtractOwnerOpptyByWeek(df_pipe: pd.DataFrame) -> Dict[str, Dict[str, int]]:
    """Extract opportunity counts per owner per week from pipe dataframe
    Counts UNIQUE opportunity numbers only (not total rows)

    Args:
        df_pipe: Pipeline DataFrame with opportunity data

    Returns:
        Dictionary mapping owner names to dictionaries of week numbers and counts
        Format: {'Owner Name': {'W01': 5, 'W02': 3, ...}, ...}
    """
    try:
        # Get column names
        owner_col = df_pipe.columns[COL_OPTYOWNER]
        created_col = df_pipe.columns[COL_CREATED]
        opty_col = df_pipe.columns[4]  # Opportunity Number column

        # Get today's date and current year for filtering
        today = datetime.now()
        current_year = today.year

        # Dictionary to store unique opportunity numbers: owner -> week -> set of opty numbers
        owner_week_opties = {}

        # Count of filtered opportunities
        future_date_count = 0
        excluded_owner_count = 0
        old_year_count = 0

        # Group by owner and created date, collecting unique opportunity numbers
        for _, row in df_pipe.iterrows():
            owner = sanitize_string_value(row[owner_col])
            created_date = row[created_col]
            opty_num = sanitize_string_value(row[opty_col])

            # Skip if owner, date, or opty number is invalid
            if owner == '' or pd.isna(created_date) or opty_num == '':
                continue

            # Skip excluded owners
            if owner in EXCLUDED_OPTY_OWNERS:
                excluded_owner_count += 1
                continue

            # Get week number from created date
            try:
                if isinstance(created_date, str):
                    created_date = pd.to_datetime(created_date, format='mixed')

                # Filter out future dates
                if created_date > today:
                    future_date_count += 1
                    logger.debug(f"Skipping future date opportunity: owner={owner}, created_date={created_date}")
                    continue

                # Filter out opportunities not from current year
                if created_date.year != current_year:
                    old_year_count += 1
                    continue

                week_num = created_date.isocalendar()[1]
                week_col = f'W{week_num:02d}'

                # Initialize owner if not exists
                if owner not in owner_week_opties:
                    owner_week_opties[owner] = {}

                # Initialize week set if not exists
                if week_col not in owner_week_opties[owner]:
                    owner_week_opties[owner][week_col] = set()

                # Add unique opportunity number to the set
                owner_week_opties[owner][week_col].add(opty_num)

            except Exception as e:
                logger.debug(f"Error processing date {created_date} for owner {owner}: {str(e)}")
                continue

        # Convert sets to counts
        owner_week_counts = {}
        for owner, weeks in owner_week_opties.items():
            owner_week_counts[owner] = {}
            for week_col, opty_set in weeks.items():
                owner_week_counts[owner][week_col] = len(opty_set)

        if future_date_count > 0:
            logger.info(f"Filtered out {future_date_count} opportunities with future creation dates")
        if excluded_owner_count > 0:
            logger.info(f"Filtered out {excluded_owner_count} opportunities from excluded owners")
        if old_year_count > 0:
            logger.info(f"Filtered out {old_year_count} opportunities from previous years (keeping only {current_year})")

        logger.info(f"Extracted unique opportunity counts for {len(owner_week_counts)} owners")
        return owner_week_counts

    except Exception as e:
        logger.error(f"Error extracting owner opportunity counts: {str(e)}")
        return {}

def ExtractOwnerOpptyDetails(df_pipe: pd.DataFrame) -> pd.DataFrame:
    """Extract detailed opportunity information for the last 5 weeks (including current week)

    Creates a table with one row per unique opportunity in the last 5 weeks.

    Args:
        df_pipe: Pipeline DataFrame with opportunity data

    Returns:
        DataFrame with columns: owner, week, details
        - owner: Opportunity owner name
        - week: "Week NN" format
        - details: Concatenated string with opty info (OpptyNum - Customer - Qty - Price - Date)
    """
    try:
        # Get column names
        owner_col = df_pipe.columns[COL_OPTYOWNER]
        created_col = df_pipe.columns[COL_CREATED]
        opty_col = df_pipe.columns[4]  # Opportunity Number column
        customer_col = df_pipe.columns[COL_CUSTOMER]
        qty_col = df_pipe.columns[7]  # Quantity column
        price_col = df_pipe.columns[COL_TOTPRICE]

        # Get today's date and current year/week for filtering
        today = datetime.now()
        current_year = today.year
        current_week = datetime.now().isocalendar()[1] if CURWEEK is None else CURWEEK

        # Calculate last 5 weeks (including current week)
        # Handle year boundary
        last_5_weeks = []
        for offset in range(-4, 1):  # -4, -3, -2, -1, 0 (last 4 weeks + current week)
            week_num = current_week + offset
            # Handle year boundaries
            if week_num < 1:
                prev_year = datetime.now().year - 1
                last_week_prev_year = datetime(prev_year, 12, 28).isocalendar()[1]
                week_num = last_week_prev_year + week_num
            elif week_num > 52:
                current_year_check = datetime.now().year
                last_week_current_year = datetime(current_year_check, 12, 28).isocalendar()[1]
                if week_num > last_week_current_year:
                    week_num = week_num - last_week_current_year
            last_5_weeks.append(week_num)

        logger.debug(f"Extracting details for last 5 weeks: {last_5_weeks}")

        # List to store detail rows
        detail_rows = []

        # Track unique opportunities per owner/week to avoid duplicates
        seen_opties = set()

        # Iterate through pipe data
        for _, row in df_pipe.iterrows():
            owner = sanitize_string_value(row[owner_col])
            created_date = row[created_col]
            opty_num = sanitize_string_value(row[opty_col])
            customer = sanitize_string_value(row[customer_col])
            qty = sanitize_numeric_value(row[qty_col])
            price = sanitize_numeric_value(row[price_col])

            # Skip if essential fields are invalid
            if owner == '' or pd.isna(created_date) or opty_num == '':
                continue

            # Skip excluded owners
            if owner in EXCLUDED_OPTY_OWNERS:
                continue

            # Process date and extract week
            try:
                if isinstance(created_date, str):
                    created_date = pd.to_datetime(created_date, format='mixed')

                # Filter out future dates
                if created_date > today:
                    continue

                # Filter out opportunities not from current year
                if created_date.year != current_year:
                    continue

                week_num = created_date.isocalendar()[1]

                # Only process if in last 5 weeks
                if week_num not in last_5_weeks:
                    continue

                # Create unique identifier to avoid duplicate rows for same opportunity
                unique_key = f"{owner}|{week_num}|{opty_num}"
                if unique_key in seen_opties:
                    continue  # Skip duplicate
                seen_opties.add(unique_key)

                # Format the details string
                # Format: OpptyNum - Customer - Qty - Price - Date
                date_str = created_date.strftime('%Y-%m-%d')
                qty_str = f"{int(qty)}" if qty > 0 else ""
                price_str = f"€{int(price):,}" if price > 0 else ""

                details = f"{opty_num} - {customer} - {qty_str} - {price_str} - {date_str}"

                # Add row to results
                detail_rows.append({
                    'owner': owner,
                    'week': f"Week {week_num:02d}",
                    'details': details
                })

            except Exception as e:
                logger.debug(f"Error processing row for details: {str(e)}")
                continue

        # Create DataFrame from collected rows
        df_details = pd.DataFrame(detail_rows, columns=['owner', 'week', 'details'])

        # Sort by owner, then week
        if not df_details.empty:
            df_details = df_details.sort_values(by=['owner', 'week']).reset_index(drop=True)

        logger.info(f"Extracted {len(df_details)} opportunity detail rows for last 5 weeks")
        return df_details

    except Exception as e:
        logger.error(f"Error extracting owner opportunity details: {str(e)}")
        return pd.DataFrame(columns=['owner', 'week', 'details'])

def UpdateOwnerOpptyTracking(df_otrack: pd.DataFrame, owner_week_counts: Dict[str, Dict[str, int]]) -> pd.DataFrame:
    """Update Owner Opportunity Tracking DataFrame with new counts, keeping maximum values

    Args:
        df_otrack: Current Owner Opportunity Tracking DataFrame
        owner_week_counts: Dictionary with new counts per owner per week

    Returns:
        Updated Owner Opportunity Tracking DataFrame
    """
    try:
        # For each owner in the new data
        for owner, week_counts in owner_week_counts.items():
            # Find existing row with this owner
            mask = df_otrack['owner'] == owner
            existing_rows = df_otrack[mask]

            if len(existing_rows) > 0:
                # Update existing row - keep maximum values
                row_idx = existing_rows.index[0]
                for week_col, new_count in week_counts.items():
                    if week_col in df_otrack.columns:
                        existing_value = pd.to_numeric(df_otrack.at[row_idx, week_col], errors='coerce')
                        if pd.isna(existing_value):
                            existing_value = 0
                        # Keep the maximum value
                        df_otrack.at[row_idx, week_col] = max(int(existing_value), new_count)
            else:
                # Create new row for this owner
                new_row = {'owner': owner}
                # Initialize all week columns to 0
                for i in range(1, 54):
                    col_name = f'W{i:02d}'
                    new_row[col_name] = week_counts.get(col_name, 0)

                # Add new row to DataFrame
                df_otrack = pd.concat([df_otrack, pd.DataFrame([new_row])], ignore_index=True)

        # Sort by owner name for consistency
        df_otrack = df_otrack.sort_values(by='owner').reset_index(drop=True)

        logger.info(f"Updated Owner Opty Tracking with {len(df_otrack)} owners")
        return df_otrack

    except Exception as e:
        logger.error(f"Error updating Owner Opty Tracking: {str(e)}")
        return df_otrack

def WriteOwnerOpptyTrackingToExcel(workbook: openpyxl.Workbook, df_otrack: pd.DataFrame, df_details: pd.DataFrame) -> None:
    """Write Owner Opportunity Tracking DataFrame to Excel, with detail table below

    Args:
        workbook: Excel workbook to write to
        df_otrack: Owner Opportunity Tracking DataFrame (Table 1) to write
        df_details: Owner Opportunity Details DataFrame (Table 2) to write

    Raises:
        PipeProcessingError: If Table 1 content exceeds the fixed start line for Table 2
    """
    try:
        # Remove existing Owner Opty Tracking tab if it exists
        if "Owner Opty Tracking" in workbook.sheetnames:
            del workbook["Owner Opty Tracking"]
            logger.debug("Removed existing Owner Opty Tracking tab")

        # Create new Owner Opty Tracking tab
        ws_otrack = workbook.create_sheet("Owner Opty Tracking")

        # Write Table 1 (summary counts) to the sheet
        for r in dataframe_to_rows(df_otrack, index=False, header=True):
            ws_otrack.append(r)

        table1_rows = len(df_otrack) + 1  # +1 for header
        logger.info(f"Written Owner Opty Tracking Table 1 with {len(df_otrack)} rows to Excel")

        # Table 2 starts at fixed line defined by LINE_LAST_5W_OPTY
        table2_start_row = LINE_LAST_5W_OPTY

        # Validate that Table 1 doesn't exceed the fixed start line for Table 2
        if table1_rows >= table2_start_row:
            error_msg = (
                f"Owner Opty Tracking Table 1 size conflict: "
                f"Table 1 has {table1_rows} rows (including header) but Table 2 must start at line {table2_start_row}. "
                f"Table 1 exceeds available space by {table1_rows - table2_start_row + 1} row(s). "
                f"Solutions: "
                f"(1) Increase LINE_LAST_5W_OPTY in .env to {table1_rows + 6} or higher, "
                f"(2) Reduce number of owners in Owner Opty Tracking, "
                f"(3) Add more owners to EXCLUDED_OPTY_OWNERS in .env"
            )
            logger.error(error_msg)
            raise PipeProcessingError(error_msg)

        # Calculate how many empty rows needed to reach the fixed start line
        empty_rows_needed = table2_start_row - table1_rows - 1  # -1 because next append is at table1_rows+1

        # Add empty rows to position Table 2 at the fixed line
        for _ in range(empty_rows_needed):
            ws_otrack.append([])

        logger.debug(f"Added {empty_rows_needed} empty rows between Table 1 and Table 2")

        # Write Table 2 (detail rows) at the fixed starting line
        for r in dataframe_to_rows(df_details, index=False, header=True):
            ws_otrack.append(r)

        logger.info(f"Written Owner Opty Tracking Table 2 with {len(df_details)} detail rows starting at fixed line {table2_start_row}")

    except PipeProcessingError:
        # Re-raise PipeProcessingError to be caught by main error handler
        raise
    except Exception as e:
        logger.error(f"Error writing Owner Opty Tracking to Excel: {str(e)}")
        raise PipeProcessingError(f"Failed to write Owner Opty Tracking: {str(e)}")

def Mapping_WeekColumn(Key: str, old_col_name: str, new_col_name: str) -> str:
    """Preserve data from existing week columns when renaming

    Args:
        Key: Unique key for the opportunity
        old_col_name: Previous column name
        new_col_name: New column name

    Returns:
        Preserved data value or empty string
    """
    try:
        # First check if the new column name already exists in master
        if new_col_name in df_master.columns:
            return Mapping_Generic(Key, new_col_name)
        # If not, check if the old column name exists and map from it
        elif old_col_name in df_master.columns:
            return Mapping_Generic(Key, old_col_name)
        else:
            return ''  # New column, no existing data
    except:
        return ''

def Format_Cell(WS: openpyxl.worksheet.worksheet.Worksheet, start: int, ColIdx: int, Format: str) -> None:
    """Apply number format to a range of cells in a worksheet

    Args:
        WS: Worksheet to format
        start: Starting row number
        ColIdx: Column index to format
        Format: Number format string
    """
    try:
        for r in range(start, WS.max_row + 1):
            WS.cell(r, ColIdx).number_format = Format
        logger.debug(f"Applied format {Format} to column {ColIdx} starting from row {start}")
    except Exception as e:
        logger.error(f"Error formatting cells: {str(e)}")

def Write2Log(wb: openpyxl.Workbook, DataLst: List[Any]) -> pd.DataFrame:
    """Write pipeline data to the log sheet

    Args:
        wb: Excel workbook to write to
        DataLst: List containing [Date, Week, Nb OPTY, Sales Force Amount, Estimated Amount]

    Returns:
        DataFrame containing the updated log data
    """

    # Check if "Pipe Log" is present
    shl = wb.sheetnames
    if "Pipe Log" not in shl:
        wslog = wb.create_sheet("Pipe Log")
        Flag = True
    else:
        wslog = wb['Pipe Log']
        Flag = False

    last_row = wslog.max_row

    if last_row == 1:
        df_log = pd.DataFrame(columns = ['Date', 'WK', 'Nb OPTY','Sales Force Amount','Estimated Amount'])
    else:
        df_log = pd.DataFrame(wslog.values)
        # Set column name from new first row
        df_log.columns = df_log.iloc[0]
        # Reset the Index
        df_log.drop(df_log.index[0],inplace=True)

    rowval = df_log.loc[df_log[GRANULARITE] == DataLst[GRANULARITE_COL]]
    if (len(rowval) != 0):
        idx = df_log.index[df_log[GRANULARITE] == DataLst[GRANULARITE_COL]]
        df_log.loc[idx] = DataLst
    else:
        df_log.loc[len(df_log)+1] = DataLst

    wslog.delete_rows(2, amount=(wslog.max_row - 1))

    for r in dataframe_to_rows(df_log, index=False, header=Flag):
        wslog.append(r)

    Format_Cell(wslog,2,1,numbers.FORMAT_DATE_DDMMYY)
    Format_Cell(wslog,2,4,'[$EUR ]#,##0_-')
    Format_Cell(wslog,2,5,'[$EUR ]#,##0_-')

    return df_log

def UpdatePipeAnalysis(wb: openpyxl.Workbook, df_log: pd.DataFrame) -> bool:
    """Update the Pipe Analysis sheet with rolling window analysis

    Args:
        wb: Excel workbook containing the analysis sheet
        df_log: DataFrame with log data for analysis

    Returns:
        Boolean indicating success/failure
    """
    # df_log expected columns:
    # 'Date', 'WK', 'Nb OPTY', 'Sales Force Amount', 'Estimated Amount'

    # Row where the Data starts (Generally 2 when the first row is used for header)
    LOGSHIFTROWDATA=2
    # Difference in row between the data start in "Log" tab versus the "analysis" tab
    # For instance 1 means that in the "Log" tab data starts row LOGSHIFTROWDATA=2 but on the
    # "Analysis" Tab it begins row 3
    SHIFTROWBETWEENTAB=1

    ret = False
    NormalizeEstimate = False

    shl = wb.sheetnames
    if "Pipe Log" not in shl:
        return ret

    # Get the Pipe Log Sheet
    wslog = wb['Pipe Log']

    if ROLLINGFIELD == 'WK':
        df_log = df_log.drop_duplicates(subset=['WK'], keep='last', ignore_index=False).copy()

    # Slicing for the last n values
    df_log = df_log.tail(ROLLINGWINDOWS).copy()
    #df_log.reset_index(inplace=True)

    # Get order of magnitude for the sales numbers
    # df_log['Magnitude'] = df_log.apply(lambda row: math.floor(math.log10(row['Sales Force Amount'])), axis = 1)
    MaxSFA = max(df_log['Sales Force Amount'])
    MinSFA = min(df_log['Sales Force Amount'])
    Mag = math.floor(math.log10(MaxSFA))

    # We substract according to its level of magnitude all common digit in the Amount serie
    # For instance if all 9 digits (Magnitude 8) amount start with 16 we substract 160000000 to the amount on the whole serie
    # The goal of the folowing loop is to find this subracted amount

    NormalizationVal = 0
    for d in range(Mag):
        df_log['Digit'] = df_log.apply(lambda row: str(row['Sales Force Amount'])[d], axis = 1)
        if len(df_log['Digit'].unique()) == 1:
            NormalizationVal = NormalizationVal + int(df_log['Digit'].unique()[0]) * 10**Mag
            Mag = Mag -1
        else:
            break

    # Check if a Normalization is also needed on the Estimated value
    # If min amount of the normalized value of Sales Force Amount is lower than the Estinate value we need to Normalize the the Estimate as well
    # For that we apply the same algorythme than before and verify that resulting difference between the 2 Normalized serie remain in the NORMAXDELTA range

    MaxSFAE = max(df_log['Estimated Amount'])
    MinSFAE = min(df_log['Estimated Amount'])
    Mag = math.floor(math.log10(MaxSFAE))

    NormalizationEVal = 0
    if MaxSFA  - NormalizationVal < MinSFAE:
        NormalizeEstimate = True
        for d in range(Mag):
            df_log['Digit'] = df_log.apply(lambda row: str(row['Estimated Amount'])[d], axis = 1)
            if len(df_log['Digit'].unique()) == 1:
                NormalizationEVal = NormalizationEVal + int(df_log['Digit'].unique()[0]) * 10**Mag
                Mag = Mag -1
            else:
                break
    # Check if Delta of Normalized value is to big (bigger than 4M)
    if (MinSFA - NormalizationVal) - (MaxSFAE - NormalizationEVal) > NORMAXDELTA:
        NormalizationVal = NormalizationVal + ((MinSFA - NormalizationVal) - (MinSFAE - NormalizationEVal) -  NORMAXDELTA)

    #Get the Pipe Log Sheet
    wsanalog = wb['Pipe Analysis']

    # Ecriture de la valeur de normalization
    # To make it more flexible la formule utilize une soustraction la valeur d'une celule fixe (R2, R=2,C=16)

    wsanalog.cell(row=2, column=18).value = NormalizationVal
    
    if NormalizeEstimate:
        wsanalog.cell(row=3, column=18).value = NormalizationEVal
    else:
        wsanalog.cell(row=3, column=18).value = 0

    # Write info on Run context
    wsanalog.cell(row=5, column=17).value = f'By {ROLLINGFIELD}'
    wsanalog.cell(row=6, column=17).value = f'For the last {ROLLINGWINDOWS} values'
    logger.info(f'Pipe Analysis with granularity on {ROLLINGFIELD}, showing the last {ROLLINGWINDOWS} records')
    # Write info on Run context
    wsanalog.cell(row=8, column=17).value = f'Last run : {date.today()}'

    wsanalog.cell(row=2, column=7).value = round(MaxSFA,0)
    wsanalog.cell(row=2, column=10).value = int(df_log['Sales Force Amount'].tail(1).iloc[0])

    #Effacement des valeurs precedentes
    for r in range(LOGSHIFTROWDATA,34):
        #Date Col = 1
        wsanalog.cell(row=(r+SHIFTROWBETWEENTAB), column=1).value = ''
        #Nb OPTY Col = 2
        wsanalog.cell(row=(r+SHIFTROWBETWEENTAB), column=2).value = ''
        #Opt Valorisation Col = 3
        wsanalog.cell(row=(r+SHIFTROWBETWEENTAB), column=3).value = ''
        #Opt Valorisation Col = 4
        wsanalog.cell(row=(r+SHIFTROWBETWEENTAB), column=4).value = ''
        #Ratio XForm Pipe Col = 16
        wsanalog.cell(row=(r+SHIFTROWBETWEENTAB), column=16).value = ''

    # Ecriture des formules dans les cellule sources du graph
    for i,r in enumerate(df_log.index):
        #Date Col = 1
        Formula = f"='Pipe Log'!A{r+1}"
        wsanalog.cell(row=(i+LOGSHIFTROWDATA+SHIFTROWBETWEENTAB), column=1).value = Formula
        #Nb OPTY Col = 2
        Formula = f"='Pipe Log'!C{r+1}"
        wsanalog.cell(row=(i+LOGSHIFTROWDATA+SHIFTROWBETWEENTAB), column=2).value = Formula
        #Opt Valorisation Col = 3
        Formula = f"='Pipe Log'!D{r+1}-$R$2"
        wsanalog.cell(row=(i+LOGSHIFTROWDATA+SHIFTROWBETWEENTAB), column=3).value = Formula
        #Opt Valorisation Col = 4
        Formula = f"='Pipe Log'!E{r+1}-$R$3"
        wsanalog.cell(row=(i+LOGSHIFTROWDATA+SHIFTROWBETWEENTAB), column=4).value = Formula
        #Ratio XForm Pipe Col = 16
        Formula = f"='Pipe Log'!E{r+1}/'Pipe Log'!D{r+1}"
        wsanalog.cell(row=(i+LOGSHIFTROWDATA+SHIFTROWBETWEENTAB), column=16).value = Formula

    # Cells Formating
    Format_Cell(wsanalog,3,1,numbers.FORMAT_DATE_DDMMYY)
    Format_Cell(wsanalog,3,3,'#,##0_-')
    Format_Cell(wsanalog,3,4,'#,##0_-')
    Format_Cell(wsanalog,3,16,'0%')
    wsanalog.cell(2,7).number_format = '[$EUR ]#,##0_-'
    wsanalog.cell(2,10).number_format = '[$EUR ]#,##0_-'

    return ret

def UpdatePipe(LatestPipe: str) -> None:
    """Main function to update pipe data with enhanced error handling"""
    global df_master, cols

    # Create colored debug message for pipe update start
    filename = os.path.basename(LatestPipe)
    colored_start_message = f"Starting pipe update process with file: {Fore.GREEN}{filename}{Style.RESET_ALL}"
    logger.debug(colored_start_message)

    try:
        # Validate input file
        if not CheckPipeFile(LatestPipe):
            raise PipeProcessingError(f"Invalid pipe file: {LatestPipe}")

        # Row where the Data starts (Generally 2 when the first row is used for header)
        HEADERSHIFT=3

        # Get creation Date for futur usage in the Log Tab
        ctimef = datetime.strptime(time.ctime(os.path.getctime(LatestPipe)), "%a %b %d %H:%M:%S %Y")

        ####################################
        # Load Latest Pipe File
        ####################################

        # Log the source directory
        source_dir_message = f'Source directory: {Fore.CYAN}{DIRECTORY_PIPE_RAW}{Style.RESET_ALL}'
        logger.info(source_dir_message)

        # Create colored log message for pipe file
        filename = os.path.basename(LatestPipe)
        colored_message = f'Using pipe file: {Fore.GREEN}{filename}{Style.RESET_ALL}'
        logger.info(colored_message)

        # Automatically detect header row (replaces manual SKIP_ROW configuration)
        try:
            skip_rows = DetectHeaderRow(LatestPipe)
            if SKIP_ROW > 0 and SKIP_ROW != skip_rows:
                logger.warning(f"SKIP_ROW={SKIP_ROW} in .env differs from auto-detected {skip_rows}. Using auto-detected value.")
                logger.warning("Note: SKIP_ROW is deprecated. Auto-detection is now used by default.")
        except Exception as e:
            logger.warning(f"Header auto-detection failed: {str(e)}. Using SKIP_ROW={SKIP_ROW}")
            skip_rows = SKIP_ROW

        # Load pipe file with auto-detected skip rows
        try:
            df_pipe = pd.read_excel(LatestPipe, skiprows=skip_rows)
            logger.info(f"Successfully loaded pipe file with {len(df_pipe)} initial rows (skipped {skip_rows} header rows)")
        except Exception as e:
            raise PipeProcessingError(f"Failed to read Excel file {LatestPipe}: {str(e)}")

        # Drop Empty Columns (more efficient with list comprehension)
        unnamed_cols = [col for col in df_pipe.columns if str(col).startswith('Unnamed:')]
        if unnamed_cols:
            df_pipe.drop(columns=unnamed_cols, inplace=True)
            logger.debug(f"Dropped {len(unnamed_cols)} unnamed columns")

        # Reorg Columns to fit the expected Master Format
        # 'Opportunity Owner','Created Date','Close Date','Stage','Opportunity Number','Indirect Account','End Customer','Estimated Quantity','Sales Price','Estimated Total Price','Sales Model Name','Part Number','Account Name','Product Line','Deal Type'
        cols = list(df_pipe.columns.values)
        # Col numbers starts at 0
        Cval = cols.pop(11)
        cols.insert(7, Cval)
        Cval = cols.pop(11)
        cols.insert(7, Cval)
        df_pipe = df_pipe.reindex(columns=cols)

        ####################################
        # Cleanup Data
        ####################################

        # Remove bogus values more efficiently with single operation
        bogus_values = ['Total', 'Confidential Information - Do Not Distribute',
                        'Copyright © 2000-2023 salesforce.com, inc. All rights reserved.']

        # Create mask for all bogus values at once
        mask = ~df_pipe[cols[COL_OPTYOWNER]].isin(bogus_values)
        df_pipe = df_pipe[mask]

        # Drop NaN values
        df_pipe.dropna(subset=[cols[COL_OPTYOWNER], cols[COL_CUSTOMER]], inplace=True)
        logger.debug(f"Removed bogus values and NaN entries")


        logger.info(f'Pipe file contains {len(df_pipe)} rows after initial processing')

    # Owner to keep
    # 'William ROMAN', 'Corinne CORDEIRO', 'Kajanan SHAN', 'Younes Giaccheri', 'Aziz ABELHAOU', 'Hippolyte FOVIAUX', 'Hatem ABBACI', 'Mehdi Dahbi', 'Gwenael BOJU', 'Charles TEZENAS', Etc ...
    
        # Owner filtering (more efficient with single isin operation)
        excluded_owners = ['Clement VIEILLEFONT', 'Vincent HALLER', 'Mathieu LUTZ', 'Calvin Chao']
        owner_mask = ~df_pipe[cols[COL_OPTYOWNER]].isin(excluded_owners)
        df_pipe = df_pipe[owner_mask]
        logger.debug(f"Filtered out excluded owners")

        # Client to Drop
        # 'Generic End User'
        # if Estimated Tot Price < 50K
        COL_SALESPRICE = 8
        COL_TOTPRICE = 9
        mask = (df_pipe[cols[COL_TOTPRICE]] < 50000) & df_pipe[cols[COL_CUSTOMER]].str.startswith('Generic')
        df_pipe.drop(df_pipe[mask].index, inplace=True)
        # Remove also the blank tot price for those 'Generic'
        mask = (df_pipe[cols[COL_TOTPRICE]]).isna() & df_pipe[cols[COL_CUSTOMER]].str.startswith('Generic')
        df_pipe.drop(df_pipe[mask].index, inplace=True)

        # Remove product lines more efficiently
        excluded_product_lines = ['LM', 'MS', 'MR']
        product_mask = ~df_pipe['Product Line'].isin(excluded_product_lines)
        df_pipe = df_pipe[product_mask]
        df_pipe['Product Line'] = df_pipe['Product Line'].fillna("")
        logger.debug(f"Filtered out excluded product lines")
        # df_pipe['Product Line'].fillna("", inplace=True) - Deprecated 3.12


        # Cleanup OPTY (remove NaN)
        df_pipe['Opportunity Number'] = df_pipe['Opportunity Number'].fillna("")
        df_pipe[cols[COL_SALESMODELNAME]] = df_pipe[cols[COL_SALESMODELNAME]].fillna("")
        # df_pipe['Opportunity Number'].fillna("", inplace=True) - Deprecated 3.12
        # df_pipe[cols[COL_SALESMODELNAME]].fillna("", inplace=True) - Deprecated 3.12

        # Format dates with error handling
        try:
            df_pipe[cols[COL_CREATED]] = pd.to_datetime(df_pipe[cols[COL_CREATED]], format='mixed', errors='coerce')
            df_pipe[cols[COL_CLOSED]] = pd.to_datetime(df_pipe[cols[COL_CLOSED]], format='mixed', errors='coerce')
            logger.debug("Date columns formatted successfully")
        except Exception as e:
            logger.warning(f"Date formatting issues: {str(e)}")

        # Copy "Run Rate" Type  Deals - But don't delete the line from the main Dataframe
        df_pipe_RR = df_pipe.loc[df_pipe['Deal Type']=='Run Rate Deal'].copy()

        # Copy "Closed Lost" Opportunities - And remove them from the main Dataframe later
        df_pipe_CL = df_pipe.loc[df_pipe[cols[COL_STAGE]]=='Closed Lost'].copy()

        # Create Key Columns (Opty+Model)
        df_pipe['Key'] = df_pipe.apply(lambda row: f'{row["Opportunity Number"]}{row[cols[COL_SALESMODELNAME]]}', axis = 1)

        logger.info(f'{len(df_pipe)} rows after data cleanup')

        ####################################
        # If Backup option is activated ... Then backup the actual Pipe file before processing.
        # Naming : name of INPUT_SUIVI_RAW "-yymmdd-hh-bck.xlsx"
        ####################################

        if BCKUP_PIPE_FILE:
            BackupPipeBefore(INPUT_SUIVI_RAW)

        ####################################
        # Load PipeLine Excel File and convert the 'Pipeline Sell Out' Tab to DataFrame
        ####################################

        try:
            myworkbook = openpyxl.load_workbook(INPUT_SUIVI_RAW, keep_vba=False)
            worksheet = myworkbook['Pipeline Sell Out']
        except Exception as e:
            raise PipeProcessingError(f"Failed to load tracking workbook {INPUT_SUIVI_RAW}: {str(e)}")

        ####################################
        # Load/Create Week History DataFrame
        ####################################

        df_whisto = LoadWeekHistoryFromExcel(myworkbook)
        logger.info(f'Week History loaded with {len(df_whisto)} rows')

        ####################################
        # Load/Create Owner Opportunity Tracking DataFrame
        ####################################

        df_otrack = LoadOwnerOpptyTrackingFromExcel(myworkbook)
        logger.info(f'Owner Opty Tracking loaded with {len(df_otrack)} rows')

        ####################################
        # Creation/Update onglet Run Rate Pipe
        ####################################

        shl = myworkbook.sheetnames
        if "Pipeline Run Rate" in shl:
            worksheet_RR= myworkbook['Pipeline Run Rate']
            # Delete all rows including header to rewrite everything
            worksheet_RR.delete_rows(1, amount=worksheet_RR.max_row)
        else:
            # Create new sheet if it doesn't exist
            worksheet_RR = myworkbook.create_sheet("Pipeline Run Rate")

        # Write DataFrame with headers
        for r in dataframe_to_rows(df_pipe_RR, index=False, header=True):
            worksheet_RR.append(r)

        ####################################
        # Creation/Update onglet Closed Lost Pipe
        ####################################

        if "Pipeline Close Lost" in shl:
            worksheet_CL= myworkbook['Pipeline Close Lost']
            # Delete all rows including header to rewrite everything
            worksheet_CL.delete_rows(1, amount=worksheet_CL.max_row)
        else:
            # Create new sheet if it doesn't exist
            worksheet_CL = myworkbook.create_sheet("Pipeline Close Lost")

        # Write DataFrame with headers
        for r in dataframe_to_rows(df_pipe_CL, index=False, header=True):
            worksheet_CL.append(r)

        ####################################
        # Version Detection and Excel Format Upgrade
        ####################################

        # Detect Excel file version based on column count
        excel_column_count = worksheet.max_column
        logger.debug(f"Detected Excel file with {excel_column_count} columns")

        if excel_column_count == V1_COLUMN_COUNT:
            logger.warning(f"Detected V1 Excel format ({V1_COLUMN_COUNT} columns). Upgrading to V2 format...")
            UpgradeFormatV1toV2(worksheet)
            excel_column_count = worksheet.max_column  # Update count after upgrade
            logger.info(f"Excel format upgraded to V2 ({excel_column_count} columns)")
        elif excel_column_count == V2_COLUMN_COUNT:
            logger.debug(f"Excel file is already V2 format ({V2_COLUMN_COUNT} columns)")
        else:
            logger.warning(f"Unexpected Excel format: {excel_column_count} columns (expected {V1_COLUMN_COUNT} or {V2_COLUMN_COUNT})")

        df_master = pd.DataFrame(worksheet.values)

        # Create colored log message for loading file
        input_filename = os.path.basename(INPUT_SUIVI_RAW)
        colored_loading_message = f'Loading Pipeline Sell Out sheet from {Fore.GREEN}{input_filename}{Style.RESET_ALL}'
        logger.info(colored_loading_message)


        # Drop first row
        df_master.drop(index=df_master.index[0], axis=0, inplace=True)

        # Set column name from new first row
        df_master.columns = df_master.iloc[0]
        # Reset the Index
        df_master = df_master.reset_index(drop=True)

        logger.info(f'Master file contains {len(df_master) - 1} rows')
        logger.info('Starting opportunity injection/refresh...')

        ####################################
        # Clean and Copy the previous value of updated columns from df_master in df_pipe when the corresponding Key (OPTY+MODEL) match
        ####################################

        # Columns de df_master
        # Common with Pipe File
        # Old Format : 'Propriétaire de l'opportunité', 'Date de création','Date de clôture', 'Etape', 'Opportunity Number', 'Revendeur','Client Final', 'Quantité', 'Prix de vente', 'Prix total','Nom du produit', 'Code du produit', 'Grossiste','Product Family\n(NX, NB, NR, PD, PT, PF)','Category Deal\nStock /CTO'
        # 'Propriétaire de l'opportunité', 'Opportunity Number', 'Date de création', 'Date de clôture', 'Étape', 'Revendeur', 'Client Final', 'Nom du produit', 'Code du produit', 'Quantité', 'Prix de vente', 'Prix total', 'Grossiste',
        # Added for Sales manual change (5 cols)
        # 'Qty\nUnit', 'Revenu projet\nK Euros', 'Quarter Invoice\nFacturation', 'Forecast projet\nMenu déroulant', 'Next Step & Support demandé / Commentaire'

        # Cleanup OPTY and Model Name (remove NaN)
        df_master['Opportunity Number'] = df_master['Opportunity Number'].fillna("")
        df_master['Nom du produit'] = df_master['Nom du produit'].fillna("")

        # Create Key Columns (Opty+Model)
        df_master['Key'] = df_master.apply(lambda row: f'{row["Opportunity Number"]}{row["Nom du produit"]}', axis = 1)
        # Master columns used for the Key while transitioning Columns Names
        #df_master['Key'] = df_master.apply(lambda row: f'{row["Date de création"]}{row["Quantité"]}', axis = 1)

        # Column Quantity
        df_pipe['Estimated\nQuantity'] = df_pipe['Key'].map(Mapping_Qty)

        # Column Revenu projet
        df_pipe['Revenu From\nEstinated Qty'] = df_pipe['Key'].map(Mapping_RevEur)

        # Column Quarter Invoice
        df_pipe['Quarter Invoice\nFacturation'] = df_pipe['Key'].map(Mapping_QtrInvoice)

        # Column Forecast projet
        # df_pipe['Forecast projet\nMenu déroulant'] = df_pipe['Key'].map(Mapping_FrCast)
        df_pipe['Forecast projet\nMenu déroulant'] = df_pipe.apply(Mapping_FrCast, axis=1)

        # Column Next Step
        df_pipe['Next Step & Support demandé / Commentaire'] = df_pipe['Key'].map(Mapping_NxtStp)

        # Dynamic Week Columns (5 columns: Week-2, Week-1, Week, Week+1, Week+2)
        dynamic_week_columns = GetDynamicWeekColumns()
        current_week = datetime.now().isocalendar()[1] if CURWEEK is None else CURWEEK
        logger.info(f'Adding dynamic week columns (current week {current_week}): {dynamic_week_columns}')

        ####################################
        # Detect week shift and copy to Week History BEFORE any updates
        ####################################

        shift_amount, existing_week_columns = DetectWeekShift(df_master)

        # Copy existing week data to Week History BEFORE any column updates
        logger.info('Copying existing week data to Week History before any shifts')
        for _, row in df_master.iterrows():
            if 'Key' in row and pd.notna(row['Key']) and str(row['Key']).strip() != '':
                key = str(row['Key'])
                # Extract Opportunity Number and Model Name
                opty_number = str(row.get('Opportunity Number', '')) if pd.notna(row.get('Opportunity Number')) else ''
                model_name = str(row.get('Nom du produit', '')) if pd.notna(row.get('Nom du produit')) else ''
                # Collect the existing week column values
                week_data = {}
                for week_col in existing_week_columns:
                    if week_col in row and pd.notna(row[week_col]) and str(row[week_col]).strip() != '':
                        week_data[week_col] = str(row[week_col])

                # Update Week History if we have any week data
                if week_data:
                    df_whisto = UpdateWeekHistoryRow(df_whisto, key, week_data, opty_number, model_name)

        ####################################
        # Apply week shift to master data if needed using history data
        ####################################

        if shift_amount != 0:
            df_master = ApplyWeekShiftFromHistory(df_master, df_whisto, dynamic_week_columns)

        ####################################
        # Update dynamic week columns with current week names
        ####################################

        # Get existing column names that might contain week data (for preservation)
        existing_week_columns = [col for col in df_master.columns if col and str(col).startswith('Week ')]

        for i, new_week_col in enumerate(dynamic_week_columns):
            # After shift, df_master has the correct data in positional columns
            if i < len(existing_week_columns):
                old_week_col = existing_week_columns[i]  # This is the positional column (Week 37, Week 38, etc.)
                # Get the shifted data from the positional column in df_master
                df_pipe[new_week_col] = df_pipe['Key'].apply(
                    lambda key: Mapping_Generic(key, old_week_col)  # Use old_week_col which has the shifted data
                )
            else:
                # New column, initialize with empty values
                df_pipe[new_week_col] = df_pipe['Key'].map(
                    lambda key: Mapping_Generic(key, new_week_col)
                )


        # Remove "Étape:Rejected"
        df_pipe.drop(df_pipe.loc[df_pipe[cols[COL_STAGE]]=='Rejected'].index, inplace=True)

        # Remove "Closed Lost" opportunities (they are now in their own tab)
        df_pipe.drop(df_pipe.loc[df_pipe[cols[COL_STAGE]]=='Closed Lost'].index, inplace=True)

        ####################################
        # Clean Week History before dropping Key column
        ####################################

        df_whisto = CleanWeekHistory(df_whisto, df_pipe)

        ####################################
        # Extract and Update Owner Opportunity Tracking
        ####################################

        logger.info('Processing Owner Opportunity Tracking data')
        owner_week_counts = ExtractOwnerOpptyByWeek(df_pipe)
        df_otrack = UpdateOwnerOpptyTracking(df_otrack, owner_week_counts)

        # Extract opportunity details for last 5 weeks
        df_opty_details = ExtractOwnerOpptyDetails(df_pipe)

        # No need of the Key Column anymore
        df_pipe.drop(['Key'], axis=1, inplace=True)
        df_master.drop(['Key'], axis=1, inplace=True)

        SFPipeAmmount = df_pipe[cols[COL_TOTPRICE]].sum()
        df_pipe['Revenu-Val'] = df_pipe['Estimated\nQuantity'].apply(pd.to_numeric, errors='coerce')
        df_pipe['Revenu-Val'] = df_pipe['Revenu-Val'] * df_pipe[cols[COL_SALESPRICE]]
        EstPipeAmmount = df_pipe['Revenu-Val'].sum()
        df_pipe.drop('Revenu-Val',axis=1,inplace=True)

        # Clean up None columns more efficiently
        try:
            none_columns = [col for col in df_master.columns if col is None]
            if none_columns:
                df_master.drop(columns=none_columns, inplace=True)
                logger.debug(f"Removed {len(none_columns)} None columns from master")
        except Exception as e:
            logger.debug(f"Error cleaning None columns: {str(e)}")

        df_pipe.columns = df_master.columns

        worksheet.delete_rows(3, amount=(worksheet.max_row - 2))

        for r in dataframe_to_rows(df_pipe, index=False, header=False):
            worksheet.append(r)

        # Update Excel column headers for dynamic Week columns (starting at column V = 22)
        # Reuse the dynamic_week_columns already calculated above
        logger.info(f'Updating Excel column headers for Week columns: {dynamic_week_columns}')
        for i, week_col_name in enumerate(dynamic_week_columns):
            col_idx = 22 + i  # V=22, W=23, X=24, Y=25, Z=26
            worksheet.cell(row=2, column=col_idx).value = week_col_name

        for i in range(HEADERSHIFT,worksheet.max_row+1):
            worksheet.cell(i,18).value = f'=Q{i}*I{i}'

        logger.info(f'Updated sheet now contains {len(df_pipe)} rows')

        # Apply Columns Formats
        # Col C = 2
        Format_Cell(worksheet,3,2,numbers.FORMAT_DATE_DDMMYY)
        # Col C = 3
        Format_Cell(worksheet,3,3,numbers.FORMAT_DATE_DDMMYY)

        # Col K = 9
        Format_Cell(worksheet,3,9,numbers.FORMAT_CURRENCY_EUR_SIMPLE)
        # Col L = 10
        Format_Cell(worksheet,3,10,'[$EUR ]#,##0_-')
        # Col Q = 17
        Format_Cell(worksheet,3,18,'[$EUR ]#,##0_-')

        # Log Pipe Data
        lst = [datetime(ctimef.year,ctimef.month,ctimef.day,0,0), ctimef.isocalendar()[1], worksheet.max_row - 2, SFPipeAmmount, EstPipeAmmount]
        logger.info(f'Updating Pipe Log with: {lst}')
        df_log = Write2Log(myworkbook,lst)

        if "Pipe Analysis" in myworkbook.sheetnames:
            logger.info('Refreshing Pipe Analysis sheet')
            UpdatePipeAnalysis(myworkbook,df_log)

        ####################################
        # Write Week History back to Excel
        ####################################

        WriteWeekHistoryToExcel(myworkbook, df_whisto)

        ####################################
        # Write Owner Opportunity Tracking back to Excel
        ####################################

        WriteOwnerOpptyTrackingToExcel(myworkbook, df_otrack, df_opty_details)

        ####################################
        # Hide configured tabs
        ####################################

        if HIDDEN_TABS:
            hidden_count = 0
            for tab_name in HIDDEN_TABS:
                if tab_name in myworkbook.sheetnames:
                    myworkbook[tab_name].sheet_state = 'hidden'
                    hidden_count += 1
                    logger.debug(f"Hidden tab: '{tab_name}'")
                else:
                    logger.warning(f"Tab '{tab_name}' not found in workbook, cannot hide")

            if hidden_count > 0:
                logger.info(f"Hidden {hidden_count} tab(s): {', '.join(HIDDEN_TABS[:3])}{'...' if len(HIDDEN_TABS) > 3 else ''}")

        myworkbook.save(OUTPUT_SUIVI_RAW)
        # Create colored log message for saving file
        output_filename = os.path.basename(OUTPUT_SUIVI_RAW)
        colored_saving_message = f'Saving to: {Fore.GREEN}{output_filename}{Style.RESET_ALL}'
        logger.info(colored_saving_message)

    except Exception as e:
        logger.error(f"Error during pipe update: {str(e)}")
        raise PipeProcessingError(f"Pipe update failed: {str(e)}")

    logger.info("Pipe update completed successfully")
    return

def main() -> None:
    """Main function with comprehensive error handling and validation"""
    try:
        logger.info("Starting UpdatePipe application")

        # Display environment configuration in DEBUG mode
        display_environment_config()

        # Validate configuration first
        validate_configuration()

        loopProc = False
        PipeFList = []

        if len(sys.argv) > 1:
            logger.info(f'Parameter {sys.argv[1]} detected')
            if sys.argv[1].lower() == 'all':
                loopProc = True
                PipeFList = GetAllPipe(DIRECTORY_PIPE_RAW)
                logger.info(f"Processing all {len(PipeFList)} pipe files")
            else:
                if CheckPipeFile(sys.argv[1]):
                    LatestPipe = sys.argv[1]
                    logger.info(f"Processing specific file: {LatestPipe}")
                else:
                    raise PipeProcessingError(f'Invalid Pipe file: {sys.argv[1]}')
        else:
            LatestPipe = GetLatestPipe(DIRECTORY_PIPE_RAW)
            # Create colored debug message for processing latest file
            filename = os.path.basename(LatestPipe)
            colored_processing_message = f"Processing latest file: {Fore.GREEN}{filename}{Style.RESET_ALL}"
            logger.debug(colored_processing_message)

        # Process files
        if loopProc:
            for i, f in enumerate(PipeFList, 1):
                logger.info(f"Processing file {i}/{len(PipeFList)}: {f}")
                UpdatePipe(f)
        else:
            UpdatePipe(LatestPipe)

        logger.info("UpdatePipe application completed successfully")

    except ConfigurationError as e:
        logger.error(f"Configuration error: {str(e)}")
        logger.error("")
        logger.error("=" * 70)
        logger.error("SETUP REQUIRED: Missing or invalid configuration")
        logger.error("=" * 70)
        logger.error("")
        logger.error("To configure this application:")
        logger.error("  1. Copy '.env.template' to '.env'")
        logger.error("     Example: copy .env.template .env")
        logger.error("")
        logger.error("  2. Edit the '.env' file and set the required variables:")
        logger.error("     - DIRECTORY_PIPE_RAW: Directory containing Salesforce reports")
        logger.error("     - INPUT_SUIVI_RAW: Path to input Excel tracking file")
        logger.error("     - OUTPUT_SUIVI_RAW: Path to output Excel tracking file")
        logger.error("")
        logger.error("  3. Run the script again")
        logger.error("")
        logger.error("=" * 70)
        sys.exit(1)
    except PipeProcessingError as e:
        logger.error(f"Processing error: {str(e)}")
        sys.exit(1)
    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main()