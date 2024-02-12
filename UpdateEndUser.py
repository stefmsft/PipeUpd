import math
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import numbers
from datetime import date,datetime
import pandas as pd
import numpy as np
import openpyxl
import glob
import os
import warnings
import sys
import time
import re
import shutil
from dotenv import load_dotenv

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

load_dotenv()

DIRECTORY_PIPE_EU_RAW = os.getenv("DIRECTORY_PIPE_EU_RAW")
INPUT_SUIVI_EU_RAW = os.getenv("INPUT_SUIVI_EU_RAW")
OUTPUT_SUIVI_EU_RAW = os.getenv("OUTPUT_SUIVI_EU_RAW")

SKIP_ROW = os.getenv("SKIP_EU_ROW")

if (SKIP_ROW == None):
    SKIP_ROW=0
else:
    SKIP_ROW = int(SKIP_ROW)

################################################################
# Functions Helper
################################################################

def GetLatestPipe(idir):

    files = glob.glob(f'{idir}/*.xls*')
    latest_file = max(files, key=os.path.getctime)

    return(latest_file)

def GetAllPipe(idir):

    files = glob.glob(f'{idir}/*.xls*')
    files.sort(key=os.path.getctime)

    return(files)


def CheckPipeFile(pfile):

    isValid = False
    if os.path.isfile(pfile):
        isValid = True
        if not ((os.path.splitext(pfile)[-1].lower() != '.xls') or (os.path.splitext(pfile)[-1].lower() != '.xlsx')):
            isValid = False

    return isValid

#Mapping Date to Quarter FYear
def GetQFFromDate(cdate):

    cm = cdate.month
    if cm < 4:
        Quarter = 1
    else:
        if cm < 7:
            Quarter = 2
        else:
            if cm < 10:
                Quarter = 3
            else:
                Quarter = 4

    Year = str(cdate.year)[-2:]

    return Quarter, Year


#Generic Mapping Functions
def Mapping_Generic  (Key,Col):
    rtv = ''

    try:
        rowval = df_master.loc[df_master['Key'] == Key]
        if (len(rowval) != 0):
            rtv = rowval.values[0][Col]
            if (rtv == None):
                rtv = ''
    except:
        pass

    return rtv

def Mapping_QtrInvoice (Key,Col,FromCol):

     # Rules :
     # If nothing, leave nothing
     # if lenght of value is not 2,4 or 6 put nothing
     # If first letter of value is Q, get the close date and calculate the QnFYyy

    eq = Mapping_Generic(Key,Col)

    seq = str(eq)

    rowpipe = df_pipe.loc[df_pipe['Key'] == Key]
    if (len(rowpipe) != 0):
        CloseDate = rowpipe.values[0][FromCol]
        if (CloseDate == None):
            CloseDate = ''
    if str(CloseDate) != '':
        Quarter,Year = GetQFFromDate(CloseDate)

    if seq == '':
            seq = f'Q{Quarter}FY{Year}'
    else:
        try:
            x = re.search("[Q]\d[F][Y]\d\d", seq)
            if None == x:
                seq = f'Q{Quarter}FY{Year}'
        except:
            pass

    return seq

def Mapping_OS (Os):
    rtv = str(Os)

    srtv = rtv.split('-')
    if len(srtv[0]) > 0:
        rtv = srtv[0]
        rtv = rtv.rstrip("\n")
        rtv = rtv.rstrip("\n1")
        rtv = rtv.rstrip("\r")

    if "academic" in rtv:
        srtv = rtv.split('*')
        if len(srtv[0]) > 0:
            rtv = srtv[0]
            rtv = rtv.rstrip("\n")
            rtv = rtv.rstrip("\r")

    return rtv


def Format_Cell(WS,start,ColIdx,Format):

    for r in range(start,WS.max_row+1):
        WS.cell(r,ColIdx).number_format = Format

    return

def UpdatePipe(LatestPipe):

    global df_master,df_pipe

    # To avoid localisation colision
    # Define col index for labels in Pipe file
    # From :
    # Opportunity Number, Quote Request Number, End Customer: Main Industry, Product: Operating System, Opportunity Owner: Full Name, Win Rate, Product Line, Created Date, Close Date, End Customer: Account Name, Indirect Account: Account Name, Account Name: Account, Name	Product: Sales Model Name, Quantity, Part Number, Requested Dealer Price, Total Price
    COL_OPTYNB=0
    COL_OS=3
    COL_OPTYOWNER=4
    COL_CREATE=7
    COL_CUSTOMER=9
    COL_SALESPN=14

    # Get creation Date for futur usage in the Log Tab
    ctimef = datetime.strptime(time.ctime(os.path.getctime(LatestPipe)), "%a %b %d %H:%M:%S %Y")

    ####################################
    # Load Latest Pipe File
    ####################################

    print(f'- Utilisation du fichier pipe : {LatestPipe}')
    # Skip SKIP_ROW if extract made with header details. Depending on the header lines this value can be updated from .env file
    df_pipe = pd.read_excel(LatestPipe, skiprows=SKIP_ROW)


    print(f'  - Il contient {len(df_pipe)} lignes')

    #######################################
    # Etape 1 : Netoyage / Reorganisation des field de l'extract pipe
    #######################################
    # From :
    # Opportunity Owner, Created Date, Close Date, Stage, Opportunity Number, Indirect Account, End Customer, Estimated Total Price, Sales Model Name, Part Number, Estimated Quantity, Sales Price, Account Name, Product Line, Deal Type
    # Target :
    # Propriétaire Opportunité, Win Rate, Close Date, Oppty N°, Disti/Sub Disti, Customer name, Project revenu, Référence produit - Modèle, P/N, Vol. oppty, PA Disti HT, Reseller, BU, Key

    ####################################
    # Cleanup Data
    ####################################

    # Drop Empty Columns if any
    for i in df_pipe.columns:
        if i.startswith('Unnamed:'):
            df_pipe = df_pipe.drop(i, axis=1)

    # Remove / Unused Values
    cols = list(df_pipe.columns.values)

    # Keep Only Kaj Team Member
    lstonw = df_pipe[cols[COL_OPTYOWNER]].unique()
    for name in ['William ROMAN', 'Corinne CORDEIRO','Kajanan SHAN','Charles TEZENAS']:
        lstonw = np.delete(lstonw, np.where(lstonw == name), axis=0)
    for name in lstonw:
        df_pipe.drop(df_pipe.loc[df_pipe[cols[COL_OPTYOWNER]]==name].index, inplace=True)

    df_pipe.dropna(subset=[cols[COL_OPTYOWNER]], inplace=True)
    df_pipe.dropna(subset=[cols[COL_CUSTOMER]], inplace=True)

    # Client to Drop
    # 'Generic End User'
    df_pipe.drop(df_pipe.loc[df_pipe[cols[COL_CUSTOMER]].str.startswith('Generic')].index, inplace=True)

    # Remove "Run Rate" Type  Deals
    # df_pipe.drop(df_pipe.loc[df_pipe['Deal Type']=='Run Rate Deal'].index, inplace=True)

    # Cleanup OPTY (remove NaN)
    df_pipe[cols[COL_OPTYNB]].fillna("", inplace=True)
    df_pipe[cols[COL_SALESPN]].fillna("", inplace=True)
    df_pipe[cols[COL_OS]].fillna("", inplace=True)

    # Cleanup OS
    df_pipe[cols[COL_OS]] = df_pipe[cols[COL_OS]].apply(lambda x: Mapping_OS(x))

    # Create Key Columns (Opty+Model)
    df_pipe['Key'] = df_pipe.apply(lambda row: f'{row[cols[COL_OPTYNB]]}{row[cols[COL_SALESPN]]}', axis = 1)

    # Delete unused Cols
    for dropcol in [COL_CREATE]:
        df_pipe.drop(cols[dropcol],axis=1,inplace=True)

    # Reaj position Close Date after removing Created
    COL_CLOSED=7
    cols = list(df_pipe.columns.values)

    # Column Quarter Invoice
    df_pipe[cols[COL_CLOSED]] = df_pipe[cols[COL_CLOSED]].apply(pd.to_datetime, format='mixed')

    # # Col numbers starts at 0
    # cols = list(df_pipe.columns.values)
    # Cval = cols.pop(len(cols)-1)
    # cols.insert(1, Cval)

    # #Reorg following the content of cols
    # df_pipe = df_pipe.reindex(columns=cols)

    # Rename Cols
    df_pipe.rename ( columns= {"End Customer: Main Industry": "Secteur",
                               "Product: Operating System": "OS",
                               "Opportunity Owner: Full Name": "Propriétaire Opportunité", \
                               "Product Line": "BU", \
                               "End Customer: Account Name": "Customer name", \
                               "Opportunity Number": "Oppty N°", \
                               "Indirect Account: Account Name": "Reseller", \
                               "Account Name: Account Name": "Disti/Sub Disti", \
                               "Quantity": "Vol. oppty", \
                               "Quote Request Number": "N° Devis",
                               "Product: Sales Model Name": "Référence produit - Modèle", \
                               "Part Number": "P/N", \
                               "Requested Dealer Price": "PA Disti HT", \
                               "Total Price": "Project revenu"}, inplace=True)


    print(f'  - {len(df_pipe)} lignes apres nettoyage')

    ####################################
    # Load PipeLine Excel File and convert the 'End Customer  Follow - up' Tab to DataFrame
    ####################################

    myworkbook=openpyxl.load_workbook(INPUT_SUIVI_EU_RAW, keep_vba=False)
    worksheet= myworkbook['End Customer  Follow - up']

    df_master = pd.DataFrame(worksheet.values)

    # Set column name from new first row
    df_master.columns = df_master.iloc[0]
    # Reset the Index
    df_master = df_master.reset_index(drop=True)

    print(f'  - Il contient {len(df_master) - 1} lignes')
    print(f'- Injection / refresh des dernieres OPTY ...')

    # Redefine the columns indirection values for step 2
    # Secteur, OS, Propriétaire Opportunité, Win Rate, BU, Periode - Invoice schedule, STATUS, Customer name, Activity, Oppty N°, IQR N°, Reseller, Disti/Sub Disti, Vol. oppty, N° Devis, Référence produit - Modèle, P/N, PA Disti HT, Project revenu,
    # Customer Capacity QTY, Customer Capacity  Value, competitors' information, Status (win/loss/commited/commited at risk/uncommited upside/uncommited), Comment, Next step, Next step schedule
    COL_OPTYNB=9
    COL_SALESPN=15
    cols = list(df_master.columns.values)

    # Cleanup OPTY and Model Name (remove NaN)
    df_master[cols[COL_OPTYNB]].fillna("", inplace=True)
    df_master[cols[COL_SALESPN]].fillna("", inplace=True)

    # Create Key Columns (Opty+Model)
    df_master['Key'] = df_master.apply(lambda row: f'{row[cols[COL_OPTYNB]]}{row[cols[COL_SALESPN]]}', axis = 1)

    #######################################
    # Etape 2 : Ajout à la fin des colonnes suplementaires à saisie manuelle
    #######################################
    # From :
    # Propriétaire Opportunité, Win Rate, Close Date, Oppty N°, Disti/Sub Disti, Customer name, Project revenu, Référence produit - Modèle, P/N, Vol. oppty, PA Disti HT, Reseller, BU, Key 
    # Target :
    # Propriétaire Opportunité, Win Rate, Oppty N°, Disti/Sub Disti, Customer name, Project revenu, Référence produit - Modèle, P/N, Vol. oppty, PA Disti HT, Reseller, BU, [ Secteur, OS, STATUS,
    # Activity, IQR N°, N° Devis, Customer Capacity QTY, Customer Capacity  Value, competitors' information, Status (win/loss/commited/commited at risk/uncommited upside/uncommited), Comment, Next step, Next step schedule, Periode - Invoice schedule ]
 
    # First map the generic columns
    # They will all appear at the end in this order
    # Indirection values from df_master from where "cols" has been set
    COL_STATUS=6 # STATUS
    COL_ACTIVITY=8 # ACTIVITY
    COL_IQR=10 # IQR
    COL_CCAPQ=19 # CCAPQ
    COL_CCAPV=20 # CCAPV
    COL_COMP=21 # COMP
    COL_STAGESTAT=22 # STAGESTAT
    COL_COMMENT=23 # COMMENT
    COL_NXT=24 # NXT
    COL_NXTD=25 # NXTD

    for extracol in [COL_STATUS,COL_ACTIVITY,COL_IQR,COL_CCAPQ,COL_CCAPV,COL_COMP,COL_STAGESTAT,COL_COMMENT,COL_NXT,COL_NXTD]:
        df_pipe[cols[extracol]] = df_pipe['Key'].apply(lambda x: Mapping_Generic(x, extracol))

    COL_PERIOD=5 # PERIOD (order in df_master)
    COL_CLOSE=7 # CLOSEDATE (order in df_pipe)
    # Column Period xform in QFY
    df_pipe[cols[COL_PERIOD]] = df_pipe['Key'].apply(lambda x: Mapping_QtrInvoice(x,COL_PERIOD,COL_CLOSE))

    # No need of the Key Column anymore
    df_pipe.drop(['Key'], axis=1, inplace=True)
    df_master.drop(['Key'], axis=1, inplace=True)

    try:
        for c in df_master:
            if None == c:
                del df_master[c]
    except:
        pass

    cols = list(df_pipe.columns.values)
    df_pipe.drop(cols[COL_CLOSE], axis=1, inplace=True)

    #######################################
    # Etape 3 : Reorg l'ordre des nouvelles colonnes ajouté
    #######################################
    # Target :
    # Secteur, OS, Propriétaire Opportunité, Win Rate, BU, Periode - Invoice schedule, STATUS, Customer name, Activity, Oppty N°, IQR N°, Reseller, Disti/Sub Disti, Vol. oppty, N° Devis, Référence produit - Modèle, P/N, PA Disti HT, Project revenu,
    # Customer Capacity QTY, Customer Capacity  Value, competitors' information, Status (win/loss/commited/commited at risk/uncommited upside/uncommited), Comment, Next step, Next step schedule

    cols = 'Secteur', 'OS', 'Propriétaire Opportunité', 'Win Rate', 'BU', 'Periode - Invoice schedule', 'STATUS', 'Customer name', 'Activity' , \
        'Oppty N°', 'IQR N°', 'Reseller', 'Disti/Sub Disti', 'N° Devis', 'Référence produit - Modèle', 'P/N', 'Vol. oppty', 'PA Disti HT', 'Project revenu', \
        'Customer Capacity QTY', 'Customer Capacity  Value', "competitors' information", 'Status (win/loss/commited/commited at risk/uncommited upside/uncommited)', 'Comment', 'Next step', 'Next step schedule'

    #Reorg following the content of cols
    df_pipe = df_pipe.reindex(columns=cols)

    df_pipe.columns = df_master.columns

    worksheet.delete_rows(2, amount=(worksheet.max_row))

    for r in dataframe_to_rows(df_pipe, index=False, header=False):
        worksheet.append(r)

    print(f'  - l onglet contient {len(df_pipe)} lignes maintenant')

    # Apply Columns Formats
    # Col S = 18
    Format_Cell(worksheet,2,18,'[$EUR ]#,##0_-')
    # Col T = 19
    Format_Cell(worksheet,2,19,'[$EUR ]#,##0_-')

    myworkbook.save(OUTPUT_SUIVI_EU_RAW)

    print(f'- Sauvegarde vers {OUTPUT_SUIVI_EU_RAW}')

    return

def main():

    loopProc = False
    PipeFList = []

    if len(sys.argv) > 1:
        print (f'Parameter {sys.argv[1]} detected')
        if sys.argv[1].lower() == 'all':
            loopProc = True
            PipeFList = GetAllPipe(DIRECTORY_PIPE_EU_RAW)
        else:
            if CheckPipeFile(sys.argv[1]):
                LatestPipe = sys.argv[1]
            else:
                print(f'Error, {sys.argv[1]} is not a valid Pipe file')
                exit()
    else:
        LatestPipe = GetLatestPipe(DIRECTORY_PIPE_EU_RAW)

    if loopProc:
        for f in PipeFList:
            UpdatePipe(f)
    else:
        UpdatePipe(LatestPipe)

    return

if __name__ == "__main__":
    main()