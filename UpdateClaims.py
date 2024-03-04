#
# Script d'integration de claims mensuel des grossistes vers un fichier de travail utilisé par UpdateEndUser.py
#
# Version : 0.1
#

import math
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import numbers
from datetime import datetime
import pandas as pd
import openpyxl
import os
import warnings
from dotenv import load_dotenv

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

load_dotenv()

DIRECTORY_PIPE_EU_RAW = os.getenv("DIRECTORY_PIPE_EU_RAW")
INPUT_SUIVI_EU_RAW = os.getenv("INPUT_SUIVI_EU_RAW")
OUTPUT_SUIVI_EU_RAW = os.getenv("OUTPUT_SUIVI_EU_RAW")

INPUT_SUIVI_EU_CLAIMH = os.getenv("INPUT_SUIVI_EU_CLAIMH")
OUTPUT_SUIVI_EU_CLAIMH = os.getenv("OUTPUT_SUIVI_EU_CLAIMH")

INPUT_CLAIM_INGRAM = os.getenv("INPUT_CLAIM_INGRAM")
INPUT_CLAIM_TD = os.getenv("INPUT_CLAIM_TD")
INPUT_CLAIM_ALSO = os.getenv("INPUT_CLAIM_ALSO")

def Format_Cell(WS,start,ColIdx,Format):

    for r in range(start,WS.max_row+1):
        WS.cell(r,ColIdx).number_format = Format

    return

def GetClaims(CProfil,SrcCol):

    df_claims = pd.DataFrame()

    print(f'- Traitement du fichier de Claim : {CProfil["Distri Name"]}')
    raw_claims = pd.read_excel(CProfil['FileName'], skiprows=CProfil['Skip'], sheet_name = CProfil['Sheet'])

    for i,c in enumerate(SrcCol):
        if c != 'Distri Name':
            df_claims[c] = raw_claims[CProfil['TrgtgCol'][i]]

    df_claims.dropna(subset=['Claim Date'], inplace=True)
    df_claims['Claim Date'] = df_claims['Claim Date'].apply(pd.to_datetime, format='mixed')
    df_claims.dropna(subset=['Quote Number'], inplace=True)
    df_claims.drop(df_claims.loc[df_claims['Quote Number'].str.startswith('Q') != True].index, inplace=True)
    df_claims.dropna(subset=['Claim Qty'], inplace=True)
    df_claims.dropna(subset=['Claim Val'], inplace=True)
    df_claims['Distri Name'] = CProfil['Distri Name']


    print(f'  - Il contient {len(df_claims)} lignes')


    return df_claims

def UpdateClaims(dfc):

    if os.path.isfile(INPUT_SUIVI_EU_CLAIMH):
        myworkbook=openpyxl.load_workbook(INPUT_SUIVI_EU_CLAIMH)
        worksheet= myworkbook['Claim History']
        df_claims = pd.DataFrame(data=worksheet.values)
        df_claims.columns = df_claims.iloc[0]
        df_claims = df_claims.reset_index(drop=True)
        df_claims.drop(index=df_claims.index[0], axis=0, inplace=True)
        worksheet.delete_rows(2, amount=(worksheet.max_row))
        FlagHeader=False
    else:
        # Create Claim History from scratch
        df_claims = pd.DataFrame(columns=list(dfc.columns.values))
        myworkbook = openpyxl.Workbook()
        worksheet = myworkbook.active
        worksheet.title = 'Claim History'
        FlagHeader=True

    new_claims = dfc[~dfc['Quote Number'].isin(df_claims['Quote Number'])]
    df_claims = pd.concat([df_claims, new_claims], ignore_index=True).copy()
    df_claims.sort_index(inplace=True)

    for r in dataframe_to_rows(df_claims, index=False, header=FlagHeader):
        worksheet.append(r)

    Format_Cell(worksheet,2,1,numbers.FORMAT_DATE_DDMMYY)

    myworkbook.save(OUTPUT_SUIVI_EU_CLAIMH)

    return

def main():

    Claims = pd.DataFrame(columns=['Claim Date','Quote Number','Claim Qty','Claim Val','Claim PN'])
    ClaimProfile = [
        {
            "Distri Name": "Ingram",
            "FileName": INPUT_CLAIM_INGRAM, 
            "Skip": 18, 
            "Sheet": 'Quotation Claim',
            "TrgtgCol": ['Invoice Date','GSN','Claim Qty','New Cost','Vendor Part No']
        },
        {
            "Distri Name": "Tech Data",
            "FileName": INPUT_CLAIM_TD, 
            "Skip": 0, 
            "Sheet": 'Claim',
            "TrgtgCol": ['Invoice date','Approval number','Qty.','Claim per pcs','Vendor Product Number']
        },
        {
            "Distri Name": "Also",
            "FileName": INPUT_CLAIM_ALSO, 
            "Skip": 0, 
            "Sheet": '107',
            "TrgtgCol": ['Invoice date','Promotion desc.','Invoice qty','Project price','Part number']
        }
        ]

    for Prf in ClaimProfile:
        Claims = pd.concat([Claims, GetClaims(Prf,list(Claims.columns.values))], ignore_index=True)

    Claims['Total'] = Claims['Claim Qty'] * Claims['Claim Val']
    Claims['Month'] = Claims['Claim Date'].apply(lambda d: d.strftime('%B'))
    print(Claims.groupby(['Distri Name','Month']).sum('Total'))

    UpdateClaims(Claims)

    return

if __name__ == "__main__":
    main()